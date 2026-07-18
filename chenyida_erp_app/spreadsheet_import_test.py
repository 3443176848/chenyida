import io
import unittest
from pathlib import Path

from openpyxl import Workbook
import xlwt

from spreadsheet_import import SpreadsheetImportError, parse_spreadsheet_import


ATTACHMENT_DIR = Path("/root/.codex/attachments/11acb435-f2b7-4453-9522-c553f1a97300")
SPEC_ATTACHMENT_DIR = Path("/root/.codex/attachments/06a0e45b-d5f3-4bef-aa0c-2dc9521194f1")


class SpreadsheetImportTest(unittest.TestCase):
    def make_xlsx(self):
        workbook = Workbook()
        cover = workbook.active
        cover.title = "说明"
        cover["A1"] = "供应商物料导入说明"
        cover["A2"] = "本页不包含物料"
        sheet = workbook.create_sheet("物料BOM")
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "物料信息"
        sheet.merge_cells("D1:F1")
        sheet["D1"] = "规格信息"
        headers = ["料号", "品名", "单位", "型号", "尺寸", "品牌"]
        for column, header in enumerate(headers, start=1):
            sheet.cell(2, column, header)
        sheet.append(["SUP-001", "贴片电阻", "PCS", "RC0603", "1.6x0.8mm", "YAGEO"])
        sheet.append(["SUP-002", "贴片电容", "PCS", "CC0402", "1.0x0.5mm", "TDK"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    def make_xls(self):
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("材料明细")
        headers = ["物料编码", "物料名称", "规格型号", "品牌", "单位"]
        for column, header in enumerate(headers):
            sheet.write(0, column, header)
        sheet.write(1, 0, "SUP-XLS-001")
        sheet.write(1, 1, "连接器")
        sheet.write(1, 2, "2.54mm 2P")
        sheet.write(1, 3, "TEST")
        sheet.write(1, 4, "PCS")
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def test_xlsx_selects_material_sheet_and_combines_multi_row_specification(self):
        result = parse_spreadsheet_import(self.make_xlsx(), "supplier.xlsx")
        self.assertEqual(result["source_type"], "XLSX")
        self.assertEqual(result["selected_sheet"], "物料BOM")
        self.assertEqual((result["header_start_row"], result["header_end_row"]), (1, 2))
        self.assertEqual(len(result["rows"]), 2)
        self.assertEqual(result["rows"][0]["raw_item_name"], "贴片电阻")
        self.assertEqual(result["rows"][0]["raw_spec"], "RC0603 1.6x0.8mm")
        self.assertEqual(result["rows"][0]["_review_status"], "NEEDS_REVIEW")
        self.assertEqual(len(result["raw_rows"]), 6)

    def test_legacy_xls_is_recognized_and_mapped(self):
        result = parse_spreadsheet_import(self.make_xls(), "supplier.xls")
        self.assertEqual(result["source_type"], "XLS")
        self.assertEqual(result["selected_sheet"], "材料明细")
        self.assertEqual(result["rows"][0]["raw_item_name"], "连接器")
        self.assertEqual(result["rows"][0]["raw_spec"], "2.54mm 2P")

    def test_csv_still_uses_same_adaptive_path(self):
        content = "标题,,,,\n物料名称,规格,品牌,单位,供应商料号\n电阻,10K 0603,YAGEO,PCS,R-1\n".encode()
        result = parse_spreadsheet_import(content, "supplier.csv")
        self.assertEqual(result["source_type"], "CSV")
        self.assertEqual(result["header_start_row"], 2)
        self.assertEqual(result["rows"][0]["raw_spec"], "10K 0603")

    def test_xlsx_content_with_csv_extension_is_detected_with_warning(self):
        result = parse_spreadsheet_import(self.make_xlsx(), "supplier.csv")
        self.assertEqual(result["source_type"], "XLSX")
        self.assertTrue(result["extension_warning"])

    def test_specification_can_be_a_review_required_name_candidate(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BOM"
        sheet.append(["物料规格描述", "物料型号", "数量"])
        sheet.append(["10K 0603", "RC0603", 1])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        result = parse_spreadsheet_import(buffer.getvalue(), "missing-name.xlsx")
        self.assertEqual(result["rows"][0]["raw_item_name"], "10K 0603")
        self.assertEqual(result["rows"][0]["_mapping_status"], "SUGGESTED")
        self.assertEqual(result["rows"][0]["_review_status"], "NEEDS_REVIEW")
        self.assertIn("MATERIAL_NAME_FROM_SPECIFICATION_REVIEW_REQUIRED", result["warnings"])

    def test_file_without_name_or_specification_semantics_fails_closed(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数据"
        sheet.append(["序号", "数量", "备注"])
        sheet.append([1, 2, "待识别"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        with self.assertRaisesRegex(SpreadsheetImportError, "物料名称列") as caught:
            parse_spreadsheet_import(buffer.getvalue(), "unmapped.xlsx")
        self.assertEqual(caught.exception.code, "IMPORT_MAPPING_REVIEW_REQUIRED")

    def test_description_only_bom_produces_review_required_specification_and_name(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BOM"
        sheet.append(["HC_CODE", "Description", "VendorCode", "Quantity", "Remark"])
        sheet.append(["HC-001", "100nF +5% 6.3V 0201", "VENDOR-001", 1, "辅助备注"])
        sheet.append(["", "", "", "", "分段标题"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = parse_spreadsheet_import(buffer.getvalue(), "description-only.xlsx")

        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["raw_item_code"], "VENDOR-001")
        self.assertEqual(result["rows"][0]["raw_item_name"], "100nF +5% 6.3V 0201")
        self.assertEqual(result["rows"][0]["raw_spec"], "100nF +5% 6.3V 0201")
        self.assertEqual(result["rows"][0]["_mapping_status"], "SUGGESTED")
        self.assertEqual(result["rows"][0]["_review_status"], "NEEDS_REVIEW")
        self.assertIn("MATERIAL_NAME_FROM_SPECIFICATION_REVIEW_REQUIRED", result["warnings"])
        self.assertEqual(result["raw_rows"][2]["disposition"], "UNMAPPED_NON_DATA")

    def test_description_column_wins_over_remark_for_specification_source(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "物料"
        sheet.append(["物料编码", "物料名称", "描述", "备注"])
        sheet.append(["M-001", "电容", "10nF 50V 0201", "仅供内部备注"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = parse_spreadsheet_import(buffer.getvalue(), "description-and-remark.xlsx")

        self.assertEqual(result["mappings"]["description"]["status"], "EXACT")
        self.assertEqual(result["mappings"]["description"]["source_headers"], ["描述"])
        self.assertEqual(result["rows"][0]["raw_spec"], "10nF 50V 0201")

    def test_model_is_kept_separate_from_rich_description_specification(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BOM"
        sheet.append(["物料名称", "物料型号", "物料描述", "生产厂家"])
        sheet.append([
            "贴片电容",
            "TEST-CAP-PART-001",
            "CAP 0201 ±5% NPO 50V 10PF TEST-CAP-PART-001",
            "测试品牌",
        ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = parse_spreadsheet_import(buffer.getvalue(), "model-and-description.xlsx")
        row = result["rows"][0]

        self.assertEqual(row["raw_model"], "TEST-CAP-PART-001")
        self.assertNotEqual(row["raw_spec"], row["raw_model"])
        self.assertIn("10PF", row["raw_spec"])
        self.assertEqual(row["_specification_evidence"], ["DESCRIPTION_STRUCTURED_SPECIFICATION_SOURCE"])

    def test_real_samples_enter_review_without_lowering_material_guards(self):
        if not ATTACHMENT_DIR.exists():
            self.skipTest("真实附件不在当前服务器")
        v700 = ATTACHMENT_DIR / "V700量产BOM.csv"
        a118 = ATTACHMENT_DIR / "A118量产BOM.csv"
        v700_result = parse_spreadsheet_import(v700.read_bytes(), v700.name)
        self.assertEqual(v700_result["selected_sheet"], "BOM")
        self.assertEqual(len(v700_result["rows"]), 222)
        self.assertTrue(all(row["_review_status"] == "NEEDS_REVIEW" for row in v700_result["rows"]))
        a118_result = parse_spreadsheet_import(a118.read_bytes(), a118.name)
        self.assertEqual(a118_result["selected_sheet"], "SHEET1")
        self.assertEqual((a118_result["header_start_row"], a118_result["header_end_row"]), (44, 44))
        self.assertEqual(len(a118_result["rows"]), 305)
        self.assertIn("SOURCE_COLUMNS_EXCEED_ANALYSIS_LIMIT_ORIGINAL_FILE_ARCHIVE_REQUIRED", a118_result["warnings"])

    def test_new_real_samples_preserve_available_specifications(self):
        if not SPEC_ATTACHMENT_DIR.exists():
            self.skipTest("规格真实附件不在当前服务器")
        expected = {
            "1928C量产BOM.xlsx": (25, 25, "M105H-DG-M23U-charge-V1_BOM"),
            "G20-G15G项目量产BOM.xlsx": (74, 69, "G9_5G_ANT_SCH_V2R0"),
            "J587_SUBA2_V01-20260703.xlsx": (122, 122, "TYPE-C耳机小板"),
        }
        for filename, (row_count, specification_count, sheet_name) in expected.items():
            with self.subTest(filename=filename):
                path = SPEC_ATTACHMENT_DIR / filename
                result = parse_spreadsheet_import(path.read_bytes(), path.name)
                self.assertEqual(result["selected_sheet"], sheet_name)
                self.assertEqual(len(result["rows"]), row_count)
                self.assertEqual(sum(bool(row["raw_spec"]) for row in result["rows"]), specification_count)
                self.assertTrue(all(row["_review_status"] == "NEEDS_REVIEW" for row in result["rows"]))
                if filename == "1928C量产BOM.xlsx":
                    capacitor = next(row for row in result["rows"] if row["_source_row_number"] == 9)
                    self.assertTrue(capacitor["raw_model"])
                    self.assertNotEqual(capacitor["raw_spec"], capacitor["raw_model"])
                    self.assertTrue(capacitor["raw_brand"])
                    self.assertTrue(capacitor["remark"])


if __name__ == "__main__":
    unittest.main()
