import csv
import io
import re
import unicodedata
import zipfile
from datetime import date, datetime
from pathlib import Path


MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_SHEETS = 50
MAX_ROWS_PER_SHEET = 50_000
MAX_TOTAL_ROWS = 100_000
MAX_COLUMNS = 256
MAX_ARCHIVE_ENTRIES = 2_048
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 80 * 1024 * 1024
HEADER_SCAN_ROWS = 50

FIELD_ALIASES = {
    "material_code": ["物料编码", "物料代码", "物料编号", "产品编码", "产品编号", "料号", "货号", "item code", "material code", "part no"],
    "material_name": ["物料名称", "物料名", "产品名称", "产品名", "品名", "名称", "货品名称", "品名及规格", "品名规格", "name", "item name", "material name"],
    "specification": ["规格", "规格型号", "型号规格", "产品规格", "规格参数", "技术规格", "规格描述", "尺寸", "参数", "品名及规格", "品名规格", "料号描述", "物料描述", "产品描述", "description", "specification", "spec", "model/spec", "size"],
    "model": ["型号", "产品型号", "物料型号", "model", "model no"],
    "brand": ["品牌", "牌子", "brand"],
    "unit": ["单位", "计量单位", "基本单位", "采购单位", "uom", "unit"],
    "category": ["分类", "物料分类", "产品分类", "类别", "品类", "category"],
    "description": ["描述", "物料描述", "产品描述", "料号描述", "说明", "备注", "description", "remark"],
    "manufacturer_part_no": ["制造商料号", "厂家料号", "厂商料号", "厂商物料编码", "厂家物料编码", "原厂料号", "mpn", "manufacturer part no"],
    "supplier_part_no": ["供应商料号", "供方料号", "供应商编码", "vendor part no", "supplier part no"],
    "drawing_no": ["图号", "图纸编号", "drawing no", "drawing number"],
    "quantity": ["数量", "用量", "需求数量", "采购数量", "qty", "quantity"],
    "price": ["单价", "价格", "含税单价", "未税单价", "price", "unit price"],
}

KEY_FIELDS = {"material_code", "material_name", "specification", "unit"}
MATERIAL_SHEET = re.compile(r"(?:^|[^a-z])bom(?:[^a-z]|$)|物料|材料|明细|清单", re.I)
NON_MATERIAL_SHEET = re.compile(r"变更|修改记录|修订|版本记录|change\s*log|revision|history", re.I)
NOTE_WORDS = re.compile(r"说明|须知|注意|报价单|报价表|价格表|目录|封面|统计|汇总|变更记录|修改记录|制表|联系人|电话|地址", re.I)
EMBEDDED_TITLE = re.compile(r"(?:^|[^a-z])bom(?:[^a-z]|$)|物料清单|材料清单", re.I)
TOTAL_WORDS = re.compile(r"^(总计|合计|累计|grand\s*total|total)$", re.I)
SUBTOTAL_WORDS = re.compile(r"^(小计|subtotal)$", re.I)
FOOTER_WORDS = re.compile(r"^(审核|批准|签字|制表|备注|页码|第\s*\d+\s*页)", re.I)
SPEC_CONTRIBUTOR = re.compile(r"型号|尺寸|材质|材料|颜色|参数|封装|等级|厚度|宽度|长度|model|size|parameter", re.I)
DIMENSION_TOKEN = re.compile(
    r"\b\d+(?:\.\d+)?\s*(?:mm|cm|m|mil|inch|英寸)?\s*[x×*]\s*"
    r"\d+(?:\.\d+)?(?:\s*[x×*]\s*\d+(?:\.\d+)?)?\s*(?:mm|cm|m|mil|inch|英寸)?\b",
    re.I,
)
TECH_TOKEN = re.compile(
    r"\b(?:0[1248]\d{2}|[A-Z]{1,6}[-/]?[A-Z0-9]{1,20}|-?\d+(?:\.\d+)?\s*"
    r"(?:kΩ|mΩ|Ω|ohm|pf|nf|uf|μf|v|kv|a|ma|w|kw|hz|mhz|ghz|℃|°c))\b",
    re.I,
)


class SpreadsheetImportError(ValueError):
    def __init__(self, message, code="IMPORT_FILE_INVALID"):
        super().__init__(message)
        self.code = code


def _text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return unicodedata.normalize("NFC", str(value)).strip()


def _normalized_header(value):
    text = unicodedata.normalize("NFKC", _text(value)).lower()
    return re.sub(r"[\s_\-—–:：;；,，.。()（）\[\]【】/\\]+", "", text)


def _detect_type(content, filename):
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xls"}:
        raise SpreadsheetImportError("仅支持 .csv、.xlsx、.xls 文件", "IMPORT_FILE_EXTENSION_UNSUPPORTED")
    if content.startswith(b"PK\x03\x04"):
        return "XLSX", suffix != ".xlsx"
    if content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
        return "XLS", suffix != ".xls"
    if suffix != ".csv":
        raise SpreadsheetImportError("文件内容与扩展名不一致", "IMPORT_FILE_SIGNATURE_MISMATCH")
    return "CSV", False


def _validate_xlsx_archive(content):
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise SpreadsheetImportError("XLSX 文件结构损坏", "IMPORT_XLSX_INVALID") from exc
    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_ARCHIVE_ENTRIES:
            raise SpreadsheetImportError("XLSX 内部文件数量超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
        names = {entry.filename.replace("\\", "/") for entry in entries}
        if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
            raise SpreadsheetImportError("文件不是有效的 XLSX 工作簿", "IMPORT_XLSX_INVALID")
        total = 0
        for entry in entries:
            name = entry.filename.replace("\\", "/")
            if name.startswith("/") or ".." in name.split("/"):
                raise SpreadsheetImportError("XLSX 包含非法内部路径", "IMPORT_XLSX_UNSAFE")
            if entry.flag_bits & 0x1:
                raise SpreadsheetImportError("不支持加密的 XLSX 文件", "IMPORT_XLSX_ENCRYPTED")
            total += entry.file_size
            if entry.compress_size and entry.file_size / entry.compress_size > 200:
                raise SpreadsheetImportError("XLSX 压缩比例超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
        if total > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
            raise SpreadsheetImportError("XLSX 解压大小超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
        lowered = {name.lower() for name in names}
        if any(name.endswith("vbaproject.bin") for name in lowered):
            raise SpreadsheetImportError("不支持包含宏的 Excel 文件", "IMPORT_XLSX_MACRO_REJECTED")
        if any(name.startswith("xl/externallinks/") for name in lowered):
            raise SpreadsheetImportError("不支持包含外部链接的 Excel 文件", "IMPORT_XLSX_EXTERNAL_LINK_REJECTED")


def _read_csv(content):
    decoded = None
    encoding = ""
    for candidate in ("utf-8-sig", "gb18030"):
        try:
            decoded = content.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise SpreadsheetImportError("CSV 编码无法识别，请使用 UTF-8 或 GB18030", "IMPORT_CSV_ENCODING_UNSUPPORTED")
    sample = decoded[:16_384]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    rows = [[_text(value) for value in row] for row in csv.reader(io.StringIO(decoded), delimiter=delimiter)]
    width = max((len(row) for row in rows), default=0)
    if len(rows) > MAX_ROWS_PER_SHEET or width > MAX_COLUMNS:
        raise SpreadsheetImportError("CSV 行数或列数超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
    return [{
        "name": "CSV",
        "rows": rows,
        "merged": [],
        "row_count": len(rows),
        "column_count": width,
    }], {"encoding": encoding, "delimiter": delimiter}


def _read_xlsx(content):
    _validate_xlsx_archive(content)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SpreadsheetImportError("服务器缺少 XLSX 解析组件", "IMPORT_XLSX_COMPONENT_MISSING") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=False, keep_links=False)
    except Exception as exc:
        raise SpreadsheetImportError("XLSX 工作簿无法解析", "IMPORT_XLSX_INVALID") from exc
    try:
        worksheets = workbook.worksheets
        if len(worksheets) > MAX_SHEETS:
            raise SpreadsheetImportError("工作表数量超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
        sheets = []
        total_rows = 0
        for worksheet in worksheets:
            row_count = worksheet.max_row or 0
            column_count = worksheet.max_column or 0
            if row_count > MAX_ROWS_PER_SHEET or column_count > MAX_COLUMNS:
                raise SpreadsheetImportError(
                    f"工作表“{worksheet.title}”的行数或列数超过安全上限",
                    "IMPORT_PARSE_LIMIT_EXCEEDED",
                )
            total_rows += row_count
            if total_rows > MAX_TOTAL_ROWS:
                raise SpreadsheetImportError("工作簿总行数超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
            rows = []
            for row in worksheet.iter_rows(min_row=1, max_row=row_count, max_col=column_count):
                rows.append([_text(cell.value) for cell in row])
            merged = [
                (cell_range.min_row, cell_range.min_col, cell_range.max_row, cell_range.max_col)
                for cell_range in worksheet.merged_cells.ranges
            ]
            sheets.append({
                "name": worksheet.title,
                "rows": rows,
                "merged": merged,
                "row_count": row_count,
                "column_count": column_count,
            })
        return sheets
    finally:
        workbook.close()


def _read_xls(content):
    try:
        import xlrd
    except ImportError as exc:
        raise SpreadsheetImportError("服务器缺少 XLS 解析组件", "IMPORT_XLS_COMPONENT_MISSING") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise SpreadsheetImportError("XLS 工作簿无法解析或已加密", "IMPORT_XLS_INVALID") from exc
    try:
        if workbook.nsheets > MAX_SHEETS:
            raise SpreadsheetImportError("工作表数量超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
        sheets = []
        total_rows = 0
        for worksheet in workbook.sheets():
            if worksheet.nrows > MAX_ROWS_PER_SHEET or worksheet.ncols > MAX_COLUMNS:
                raise SpreadsheetImportError(
                    f"工作表“{worksheet.name}”的行数或列数超过安全上限",
                    "IMPORT_PARSE_LIMIT_EXCEEDED",
                )
            total_rows += worksheet.nrows
            if total_rows > MAX_TOTAL_ROWS:
                raise SpreadsheetImportError("工作簿总行数超过安全上限", "IMPORT_PARSE_LIMIT_EXCEEDED")
            rows = [
                [_text(worksheet.cell_value(row_index, column_index)) for column_index in range(worksheet.ncols)]
                for row_index in range(worksheet.nrows)
            ]
            merged = [(rlo + 1, clo + 1, rhi, chi) for rlo, rhi, clo, chi in worksheet.merged_cells]
            sheets.append({
                "name": worksheet.name,
                "rows": rows,
                "merged": merged,
                "row_count": worksheet.nrows,
                "column_count": worksheet.ncols,
            })
        return sheets
    finally:
        workbook.release_resources()


def _propagated_header(sheet, start, end):
    propagated = {}
    for row_number in range(start, end + 1):
        row = sheet["rows"][row_number - 1] if row_number <= len(sheet["rows"]) else []
        for column_index, value in enumerate(row, start=1):
            if value:
                propagated[(row_number, column_index)] = value
    for min_row, min_column, max_row, max_column in sheet["merged"]:
        if min_row < start or min_row > end:
            continue
        parent = propagated.get((min_row, min_column), "")
        if not parent:
            continue
        for row_number in range(min_row, min(max_row, end) + 1):
            for column_index in range(min_column, max_column + 1):
                propagated.setdefault((row_number, column_index), parent)
    return propagated


def _header_columns(sheet, start, end):
    propagated = _propagated_header(sheet, start, end)
    columns = []
    for column_index in range(1, sheet["column_count"] + 1):
        parts = []
        for row_number in range(start, end + 1):
            part = propagated.get((row_number, column_index), "")
            if part and (not parts or parts[-1] != part):
                parts.append(part)
        samples = []
        for row in sheet["rows"][end:end + 20]:
            value = row[column_index - 1] if column_index <= len(row) else ""
            if value:
                samples.append(value)
        numeric = sum(bool(re.fullmatch(r"-?(?:\d+|\d+\.\d+)", value)) for value in samples)
        columns.append({
            "index": column_index - 1,
            "header": "/".join(parts),
            "samples": samples[:10],
            "numeric_ratio": numeric / len(samples) if samples else 0,
            "unique_ratio": len(set(samples)) / len(samples) if samples else 0,
        })
    return columns


def _matching_fields(header):
    normalized = _normalized_header(header)
    result = set()
    if not normalized:
        return result
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            target = _normalized_header(alias)
            if normalized == target or normalized.endswith(target) or (len(target) >= 3 and target in normalized):
                if field == "material_code" and re.search(r"厂商|厂家|制造商|供应商|供方|原厂", normalized):
                    continue
                result.add(field)
    return result


def _header_candidate(sheet, start, span):
    end = start + span - 1
    columns = _header_columns(sheet, start, end)
    headers = [column["header"] for column in columns if column["header"]]
    matched = [_matching_fields(header) for header in headers]
    alias_hits = sum(bool(fields) for fields in matched)
    fields = set().union(*matched) if matched else set()
    key_hits = len(fields & KEY_FIELDS)
    ratio = len(headers) / max(1, sheet["column_count"])
    uniqueness = len({_normalized_header(header) for header in headers}) / max(1, len(headers))
    following = sheet["rows"][end:end + 10]
    widths = [sum(bool(value) for value in row) for row in following if any(row)]
    stability = 0
    if widths:
        stability = 0.5 if len(widths) == 1 else 1 - (max(widths) - min(widths)) / max(1, max(widths))
    substantive = sum(width >= max(2, (sheet["column_count"] + 1) // 2) for width in widths)
    unmerged_single_cell_preambles = 0
    for row_number in range(start, end):
        row = sheet["rows"][row_number - 1] if row_number <= len(sheet["rows"]) else []
        populated = [value for value in row if value]
        if len(populated) != 1 or sheet["column_count"] < 2:
            continue
        covered = any(
            min_row == row_number and max_column > min_column
            for min_row, min_column, _max_row, max_column in sheet["merged"]
        )
        if not covered:
            unmerged_single_cell_preambles += 1
    score = (
        ratio * 0.15
        + uniqueness * 0.10
        + min(1, alias_hits / 4) * 0.32
        + min(1, key_hits / 3) * 0.20
        + stability * 0.10
        + min(1, substantive / 2) * 0.13
    )
    if not substantive:
        score -= 0.20
    score -= unmerged_single_cell_preambles * 0.35
    return {
        "start": start,
        "end": end,
        "data_start": end + 1,
        "score": max(0, min(1, score)),
        "alias_hits": alias_hits,
        "key_hits": key_hits,
        "columns": columns,
    }


def _analyze_sheet(sheet, index):
    candidates = []
    scan_end = min(HEADER_SCAN_ROWS, sheet["row_count"])
    for start in range(1, scan_end + 1):
        for span in range(1, 4):
            if start + span - 1 <= scan_end:
                candidates.append(_header_candidate(sheet, start, span))
    candidates.sort(key=lambda item: (-item["score"], -item["alias_hits"], item["start"], item["end"]))
    header = candidates[0] if candidates and candidates[0]["score"] >= 0.35 else None
    non_empty_rows = sum(any(row) for row in sheet["rows"])
    non_empty_columns = len({
        column_index
        for row in sheet["rows"]
        for column_index, value in enumerate(row)
        if value
    })
    continuity = non_empty_rows / max(1, sheet["row_count"])
    score = (
        (header["score"] if header else 0) * 0.65
        + min(1, non_empty_rows / 10) * 0.15
        + min(1, non_empty_columns / 4) * 0.10
        + continuity * 0.10
    )
    if MATERIAL_SHEET.search(sheet["name"]):
        score += 0.12
    if NON_MATERIAL_SHEET.search(sheet["name"]):
        score -= 0.28
    return {
        "index": index,
        "name": sheet["name"],
        "score": max(0, min(1, score)),
        "header": header,
    }


def _mapping_score(field, column):
    normalized = _normalized_header(column["header"])
    best = 0
    exact = False
    for alias in FIELD_ALIASES[field]:
        target = _normalized_header(alias)
        if not target:
            continue
        if normalized == target or normalized.endswith(target):
            best = 1
            exact = True
            break
        if len(target) >= 3 and target in normalized:
            best = max(best, 0.82)
    if field == "material_code" and re.search(r"厂商|厂家|制造商|供应商|供方|原厂", normalized):
        return 0, False
    samples = column["samples"]
    sample_score = 0
    if samples:
        if field in {"quantity", "price"}:
            sample_score = column["numeric_ratio"] * 0.8
        elif field == "unit":
            sample_score = sum(bool(re.fullmatch(r"(?:pcs?|个|件|套|卷|米|m|kg|g|箱|包|片)", value, re.I)) for value in samples) / len(samples) * 0.8
        elif field in {"specification", "model"}:
            sample_score = sum(bool(DIMENSION_TOKEN.search(value) or TECH_TOKEN.search(value) or re.search(r"[A-Za-z].*\d|\d.*[A-Za-z]", value)) for value in samples) / len(samples) * 0.7
        elif field in {"material_code", "supplier_part_no", "manufacturer_part_no"}:
            sample_score = sum(bool(re.search(r"\d", value) and not re.search(r"\s", value)) for value in samples) / len(samples) * 0.5 + column["unique_ratio"] * 0.3
        else:
            sample_score = column["unique_ratio"] * 0.2
    return min(1, best * 0.78 + sample_score * 0.20), exact


def _suggest_mappings(header):
    mappings = {}
    for field in FIELD_ALIASES:
        scored = []
        for column in header["columns"]:
            score, exact = _mapping_score(field, column)
            if score >= 0.35:
                scored.append((score, column["index"], exact, column["header"]))
        scored.sort(key=lambda item: (-item[0], item[1]))
        selected = scored[:1]
        strategy = "FIRST_NON_EMPTY"
        if field == "specification":
            dedicated = [
                item
                for item in scored
                if item[2]
                and re.fullmatch(
                    r"规格|规格型号|型号规格|产品规格|规格参数|技术规格|规格描述|specification|spec|model/spec",
                    item[3].split("/")[-1],
                    re.I,
                )
            ]
            contributors = []
            for column in header["columns"]:
                if SPEC_CONTRIBUTOR.search(column["header"]):
                    score, exact = _mapping_score(field, column)
                    contributors.append((max(score, 0.72), column["index"], exact, column["header"]))
            contributors.sort(key=lambda item: (-item[0], item[1]))
            selected = dedicated[:1] if dedicated else contributors[:5]
            selected.sort(key=lambda item: item[1])
            strategy = "JOIN_NON_EMPTY" if len(selected) > 1 else "SPECIFICATION_EXTRACT"
        if not selected:
            mappings[field] = {"indexes": [], "headers": [], "status": "UNMAPPED", "confidence": 0, "strategy": strategy}
            continue
        confidence = sum(item[0] for item in selected) / len(selected)
        conflict = field != "specification" and len(scored) > 1 and abs(scored[0][0] - scored[1][0]) <= 0.04
        status = "CONFLICT" if conflict else "EXACT" if selected[0][2] and len(selected) == 1 else "HIGH_CONFIDENCE" if confidence >= 0.82 else "SUGGESTED"
        mappings[field] = {
            "indexes": [item[1] for item in selected],
            "headers": [item[3] for item in selected],
            "status": status,
            "confidence": round(confidence, 4),
            "strategy": strategy,
        }
    return mappings


def _classify_row(row, header, row_number):
    values = [value for value in row if value]
    if not values:
        return "BLANK"
    joined = " ".join(values).strip()
    if TOTAL_WORDS.fullmatch(joined):
        return "TOTAL"
    if SUBTOTAL_WORDS.fullmatch(joined):
        return "SUBTOTAL"
    if FOOTER_WORDS.search(joined):
        return "FOOTER"
    if header["start"] <= row_number <= header["end"]:
        return "TITLE_OR_NOTE"
    header_leaves = {_normalized_header(column["header"].split("/")[-1]) for column in header["columns"] if column["header"]}
    matches = sum(_normalized_header(value) in header_leaves for value in values)
    if row_number > header["end"] and matches >= min(2, max(1, len(header_leaves))):
        return "REPEATED_HEADER"
    if row_number < header["start"] and (len(values) == 1 or NOTE_WORDS.search(joined)):
        return "TITLE_OR_NOTE"
    if len(values) <= 2 and (NOTE_WORDS.search(joined) or EMBEDDED_TITLE.search(joined)):
        return "TITLE_OR_NOTE"
    return "DATA"


def _resolve(row, mapping):
    values = [row[index] if index < len(row) else "" for index in mapping["indexes"]]
    values = list(dict.fromkeys(value for value in values if value and not value.lstrip().startswith("=")))
    if not values or mapping["status"] in {"UNMAPPED", "CONFLICT"}:
        return ""
    return values[0] if mapping["strategy"] == "FIRST_NON_EMPTY" else " ".join(values)


def _extract_specification(row, mappings, material_name, description):
    mapping = mappings["specification"]
    explicit = _resolve(row, mapping)
    if explicit:
        return explicit, mapping["confidence"], mapping["status"], ["EXPLICIT_OR_COMPONENT_SPECIFICATION"]
    fallback = " ".join(value for value in (material_name, description) if value)
    tokens = list(dict.fromkeys(DIMENSION_TOKEN.findall(fallback) + TECH_TOKEN.findall(fallback)))
    if not tokens:
        return "", 0, "UNMAPPED", ["NO_SPECIFICATION_EVIDENCE"]
    return " ".join(tokens), min(0.68, 0.48 + len(tokens) * 0.05), "SUGGESTED", ["DETERMINISTIC_NAME_OR_DESCRIPTION_EXTRACTION", "MANUAL_CONFIRMATION_REQUIRED"]


def parse_spreadsheet_import(content, filename):
    if not content:
        raise SpreadsheetImportError("导入文件不能为空", "IMPORT_FILE_EMPTY")
    if len(content) > MAX_FILE_BYTES:
        raise SpreadsheetImportError("导入文件不能超过 10 MiB", "IMPORT_FILE_TOO_LARGE")
    source_type, extension_warning = _detect_type(content, filename)
    csv_meta = {}
    if source_type == "CSV":
        sheets, csv_meta = _read_csv(content)
    elif source_type == "XLSX":
        sheets = _read_xlsx(content)
    else:
        sheets = _read_xls(content)
    if not sheets:
        raise SpreadsheetImportError("文件中没有可读取的工作表", "IMPORT_NO_WORKSHEET")
    analyses = [_analyze_sheet(sheet, index) for index, sheet in enumerate(sheets)]
    analyses.sort(key=lambda item: (-item["score"], item["index"]))
    selected_analysis = analyses[0]
    header = selected_analysis["header"]
    if not header:
        raise SpreadsheetImportError("无法识别物料表头，请人工整理或另存后重试", "IMPORT_HEADER_REVIEW_REQUIRED")
    selected_sheet = sheets[selected_analysis["index"]]
    mappings = _suggest_mappings(header)
    if mappings["material_name"]["status"] in {"UNMAPPED", "CONFLICT"}:
        raise SpreadsheetImportError("未识别到明确的物料名称列，已停止导入以避免错误建档", "IMPORT_MAPPING_REVIEW_REQUIRED")
    import_rows = []
    raw_rows = []
    for sheet in sheets:
        for row_number, row in enumerate(sheet["rows"], start=1):
            raw_rows.append({
                "sheet_name": sheet["name"],
                "row_number": row_number,
                "values": row,
                "disposition": "UNSELECTED_SHEET" if sheet is not selected_sheet else _classify_row(row, header, row_number),
            })
    for raw in raw_rows:
        if raw["sheet_name"] != selected_sheet["name"] or raw["disposition"] != "DATA":
            continue
        row = raw["values"]
        mapped = {field: _resolve(row, mapping) for field, mapping in mappings.items()}
        specification, specification_confidence, specification_status, spec_evidence = _extract_specification(
            row,
            mappings,
            mapped["material_name"],
            mapped["description"],
        )
        mapped["specification"] = specification
        key_statuses = [mappings["material_name"]["status"], specification_status]
        mapping_status = "CONFLICT" if "CONFLICT" in key_statuses else "UNMAPPED" if "UNMAPPED" in key_statuses else "SUGGESTED" if "SUGGESTED" in key_statuses else "HIGH_CONFIDENCE" if "HIGH_CONFIDENCE" in key_statuses else "EXACT"
        mapping_confidence = round((mappings["material_name"]["confidence"] + specification_confidence) / 2, 4)
        import_rows.append({
            "raw_item_name": mapped["material_name"],
            "raw_item_code": mapped["supplier_part_no"] or mapped["material_code"],
            "raw_spec": specification,
            "raw_brand": mapped["brand"],
            "raw_mpn": mapped["manufacturer_part_no"],
            "purchase_uom": mapped["unit"],
            "last_price": mapped["price"],
            "remark": mapped["description"],
            "_source_sheet_name": selected_sheet["name"],
            "_source_row_number": raw["row_number"],
            "_mapped_values": mapped,
            "_mapping_confidence": mapping_confidence,
            "_specification_confidence": round(specification_confidence, 4),
            "_mapping_status": mapping_status,
            "_review_status": "AUTO_ACCEPTABLE" if mapping_status == "EXACT" and specification_status == "EXACT" else "NEEDS_REVIEW",
            "_specification_evidence": spec_evidence,
        })
    if not import_rows:
        raise SpreadsheetImportError("未识别到可导入的物料数据行", "IMPORT_NO_DATA_ROWS")
    mapping_summary = {
        field: {
            "source_headers": mapping["headers"],
            "status": mapping["status"],
            "confidence": mapping["confidence"],
            "strategy": mapping["strategy"],
        }
        for field, mapping in mappings.items()
    }
    return {
        "source_type": source_type,
        "extension_warning": extension_warning,
        "selected_sheet": selected_sheet["name"],
        "header_start_row": header["start"],
        "header_end_row": header["end"],
        "data_start_row": header["data_start"],
        "structure_confidence": round(selected_analysis["score"], 4),
        "columns": [column["header"] for column in header["columns"]],
        "mappings": mapping_summary,
        "rows": import_rows,
        "raw_rows": raw_rows,
        "sheet_summaries": [
            {
                "name": sheet["name"],
                "row_count": sheet["row_count"],
                "column_count": sheet["column_count"],
            }
            for sheet in sheets
        ],
        **csv_meta,
    }
