import importlib.util
import json
import os
import stat
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


MODULE_PATH = Path(__file__).parents[1] / "tools" / "real-bom-import" / "export-standardized-bom-workbook.py"
SPEC = importlib.util.spec_from_file_location("standardized_bom_workbook", MODULE_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


def make_template(path: Path) -> None:
    workbook = openpyxl.Workbook()
    raw = workbook.active
    raw.title = "原BOM"
    raw.append(["A200-USB-BOM 8SD05169C"])
    raw.append(["序号", "物料规格描述", "物料型号", "数量", "位号", "备注", "替代料"])
    target_materials = []
    for number in range(1, 56):
        if number in (20, 40):
            raw.append([number, f"PCB,A200-BOARD-{number}", f"A200-BOARD-{number}", 1, None, None, None])
            continue
        model = f"TEST-MPN-{number:03d}"
        quantity = number % 5 + 1
        raw.append([number, f"IC,{model},TESTBRAND", model, quantity, f"U{number}", None, None])
        target_materials.append((model, quantity))
    target = workbook.create_sheet("Sheet1")
    target.append(list(MODULE.TEMPLATE_HEADERS))
    for sequence, (model, quantity) in enumerate(target_materials, 1):
        target.append(
            [
                sequence,
                "A200",
                "USB小板",
                "8SD05169C",
                model,
                "TESTBRAND",
                quantity,
                None,
                "测试供应商",
                None,
                f"=G{sequence + 1}*J{sequence + 1}",
                f"=K{sequence + 1}-M{sequence + 1}",
                None,
            ]
        )
    workbook.save(path)
    workbook.close()


class StandardizedBomWorkbookTest(unittest.TestCase):
    def test_template_pair_validates_all_53_groups(self):
        with tempfile.TemporaryDirectory() as directory:
            template = Path(directory) / "moban.xlsx"
            make_template(template)
            result = MODULE.validate_template_pair(template)
        self.assertEqual(result["raw_groups"], 53)
        self.assertEqual(result["target_rows"], 53)
        self.assertEqual(result["raw_board_rows_excluded"], 2)
        self.assertEqual(result["row_evidence_matches"], 53)
        self.assertEqual(result["quantity_matches"], 53)

    def test_classic_parser_groups_alternatives_and_excludes_board(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "G20-G15G项目量产BOM.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "G9_5G_ANT_SCH_V2R0"
            sheet.append(["G20-5G小板（4SH05429A）"])
            sheet.append(["Item", "HC_CODE", "Description", "VendorCode", "Vendor", "Quantity", "Reference", "替代料"])
            sheet.append([1, "HC1", "RES-CHIP 0.05w-0ohm-0201", "MAIN-001", "Maker A", 2, "R1,R2", None])
            sheet.append([None, "HC1B", "RES-CHIP 0.05w-0ohm-0201", "ALT-001", "Maker B", None, None, None])
            sheet.append([2, "HC2", "PCB,G20-SUB-V1", "G20-SUB-V1", "Board", 1, "PCB", None])
            workbook.save(source)
            workbook.close()
            groups, anomalies, stats = MODULE.parse_classic_groups(
                source,
                project="G20-G15G",
                sheet_name="G9_5G_ANT_SCH_V2R0",
            )
            rows = MODULE.standard_rows_from_groups(groups)
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0].quantity, 2)
        self.assertEqual([line.row_no for line in groups[0].alternatives], [4])
        self.assertEqual(rows[0].board_type, "5G小板")
        self.assertEqual(rows[0].internal_model, "4SH05429A")
        self.assertIn("电阻", rows[0].specification)
        self.assertEqual(rows[0].substitute, "ALT-001(Maker B)")
        self.assertEqual(stats["board_base_excluded"], 1)
        self.assertEqual([item["kind"] for item in anomalies], ["板件本体未计入"])

    def test_j587_unlabelled_quantity_column_is_explicitly_mapped(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "J587_SUBA2_V01-20260703.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "TYPE-C耳机小板"
            sheet.append(["J587_SUBA1_V01_20251215"])
            sheet.append(["序号", "状态", "物料编码", "物料名称", "描述", "品牌", "供应商料号", "位号", None, "备注"])
            sheet.append([1, "主料", "02.1", "Resistor", "RES,10K,±1%,1/20W,0201", "UniOhm", "MAIN", "R1 R2 R3", 3, None])
            sheet.append([None, "替代料", "02.1.1", "Resistor", "RES,10K,±1%,1/20W,0201", "Yageo", "ALT", None, None, None])
            workbook.save(source)
            workbook.close()
            result = MODULE.parse_source(
                source,
                project="J587",
                sheet_name="TYPE-C耳机小板",
                special_j587=True,
            )
        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].quantity, 3)
        self.assertEqual(result.rows[0].internal_model, "J587_SUBA1_V01")
        self.assertEqual(result.rows[0].substitute, "ALT(Yageo)")

    def test_a200_material_list_keeps_only_explicit_supplier_and_unknown_quantity(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "A200量产物料清单.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet2"
            sheet.append(["A200项目BOM表清单及成本核算"])
            sheet.append([])
            sheet.append(["编号", "物料名称", "物料编码", "物料规格", "供应商"])
            sheet.append(["一", "8层HDI-USB小板", "8SD05169C", "板厚0.6", "板厂"])
            sheet.append([1, "钢网", None, "0.1MM", "联创佳"])
            sheet.append([2, "包装", None, "100/包", "晨亿达"])
            workbook.save(source)
            workbook.close()
            result = MODULE.parse_a200_material_list(source)
        self.assertEqual(len(result.rows), 2)
        self.assertEqual(result.rows[0].board_type, "USB小板")
        self.assertEqual(result.rows[0].internal_model, "8SD05169C")
        self.assertEqual(result.rows[0].supplier, "联创佳")
        self.assertIsNone(result.rows[0].quantity)
        self.assertEqual(len(result.anomalies), 2)

    def test_identical_repeated_bom_section_is_kept_once(self):
        context_a = MODULE.Context("A118", "DMR/LTE小板", "4SH04126A", "A118_DMR_BOM 4SH04126A")
        first = MODULE.RawGroup(
            primary=MODULE.RawLine(
                filename="A118量产BOM.xlsx",
                sheet="SHEET1",
                row_no=10,
                context=context_a,
                name="电阻",
                specification="0201,10K,±1%",
                manufacturer_part_no="MAIN",
                quantity_raw=1,
                reference="R1",
            ),
            quantity=1,
            quantity_source="原表用量",
        )
        context_b = MODULE.Context("A118", "DMR/LTE小板", "4SH04126A", "A118_DMR_BOM 4SH04126A")
        repeated = MODULE.RawGroup(
            primary=MODULE.RawLine(
                filename="A118量产BOM.xlsx",
                sheet="SHEET1",
                row_no=100,
                context=context_b,
                name="电阻",
                specification="0201,10K,±1%",
                manufacturer_part_no="MAIN",
                quantity_raw=1,
                reference="R1",
            ),
            quantity=1,
            quantity_source="原表用量",
        )
        separating = MODULE.RawGroup(
            primary=MODULE.RawLine(
                filename="A118量产BOM.xlsx",
                sheet="SHEET1",
                row_no=50,
                context=MODULE.Context("A118", "USB小板", "4SD00001A", "USB 4SD00001A"),
                name="IC",
                specification="USB-IC",
                manufacturer_part_no="USB-IC",
                quantity_raw=1,
                reference="U1",
            ),
            quantity=1,
            quantity_source="原表用量",
        )
        kept, anomalies, stats = MODULE.suppress_identical_repeated_sections([first, separating, repeated])
        self.assertEqual([group.primary.row_no for group in kept], [10, 50])
        self.assertEqual(stats["duplicate_sections_suppressed"], 1)
        self.assertEqual(stats["duplicate_groups_suppressed"], 1)
        self.assertEqual(anomalies[0]["kind"], "重复BOM区段未重复汇总")

    def test_workbook_contract_formulas_and_provenance(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "moban.xlsx"
            output = root / "result.xlsx"
            make_template(template)
            standard_row = MODULE.StandardRow(
                project="TEST",
                board_type="测试板",
                internal_model="2SH00001A",
                specification="电阻,10K,±1%,0201",
                brand="Brand",
                quantity=2,
                substitute="ALT",
                supplier="Supplier",
                source_file="source.xlsx",
                source_sheet="BOM",
                source_row=3,
                alternative_rows=(4,),
                rule="合成测试",
            )
            sheets = {title: [] for title in MODULE.SOURCE_SHEET_ORDER}
            sheets["1928C标准"] = [standard_row]
            data = {
                "sheets": sheets,
                "consolidated": [standard_row],
                "anomalies": [],
                "source_stats": {
                    filename: {"handling": "PARSED_TO_STANDARD", "standard_rows": 0}
                    for filename in MODULE.EXPECTED_SOURCE_FILES
                },
                "template_validation": {"target_rows": 53},
                "counts": {
                    "standard_rows": 1,
                    "source_standard_sheets": 8,
                    "nonempty_source_standard_sheets": 1,
                    "anomalies": 0,
                    "unknown_quantity_rows": 0,
                    "alternative_bearing_rows": 1,
                },
            }
            data["source_stats"]["A200量产BOM.xlsx"]["handling"] = "SUPERSEDED_BY_MOBAN_CANONICAL_PAIR"
            data["source_stats"]["A200量产物料清单.xlsx"]["handling"] = "PARSED_TO_STANDARD_WITH_UNKNOWN_QUANTITY"
            data["source_stats"]["A200量产注意事项.xls"]["handling"] = "ARCHIVE_NOTE_NO_MATERIAL_ROWS"
            manifest = {
                "files": [
                    {"filename": filename, "sha256": "0" * 64}
                    for filename in sorted(MODULE.EXPECTED_SOURCE_FILES)
                ]
            }
            template_snapshot = {"filename": "moban.xlsx", "sha256": "1" * 64}
            MODULE.write_workbook(output, data, template, manifest, template_snapshot)
            validation = MODULE.validate_workbook(output, data)
            mode = stat.S_IMODE(output.stat().st_mode)
            workbook = openpyxl.load_workbook(output, data_only=False, keep_links=False)
            try:
                consolidated = workbook["全部物料汇总"]
                self.assertEqual(tuple(consolidated.cell(1, column).value for column in range(1, 14)), MODULE.TEMPLATE_HEADERS)
                self.assertEqual(consolidated["K2"].value, '=IF(OR(G2="",J2=""),"",G2*J2)')
                self.assertEqual(consolidated["L2"].value, '=IF(OR(K2="",M2=""),"",MAX(K2-M2,0))')
                self.assertIn("source.xlsx", consolidated["A2"].comment.text)
                notes = "\n".join(
                    str(cell.value or "")
                    for row in workbook["来源与说明"].iter_rows()
                    for cell in row
                )
                self.assertIn("更新记录、变更记录及空Sheet3", notes)
            finally:
                workbook.close()
        self.assertEqual(mode, 0o600)
        self.assertEqual(validation["standard_rows"], 1)
        self.assertEqual(validation["formula_cells"], 4)

    def test_formula_like_source_text_is_neutralized(self):
        self.assertEqual(MODULE.safe_excel_text("=HYPERLINK('x')"), "'=HYPERLINK('x')")
        self.assertEqual(MODULE.safe_excel_text("+1+1"), "'+1+1")


if __name__ == "__main__":
    unittest.main()
