import csv
import hashlib
import importlib.util
import json
import os
import stat
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


MODULE_PATH = Path(__file__).parents[1] / "tools" / "real-bom-import" / "export-internal-library.py"
SPEC = importlib.util.spec_from_file_location("internal_material_library_export", MODULE_PATH)
exporter = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = exporter
SPEC.loader.exec_module(exporter)


class InternalMaterialLibraryExportTest(unittest.TestCase):
    def test_reference_quantity_is_conservative(self):
        self.assertEqual(exporter.explicit_reference_quantity("R1,R2 R3"), 3)
        self.assertIsNone(exporter.explicit_reference_quantity("R1-R3"))
        self.assertIsNone(exporter.explicit_reference_quantity("R1,R1"))

    def test_template_match_requires_exact_normalized_identity_text(self):
        template = exporter.TemplateRow(2, "A200", "USB小板", "8SD00001A", "ABC-123", "品牌", 1, "", "", None, None)
        sources = [
            {"specification": "ABC-1234", "model": "", "manufacturer_part_no": ""},
            {"specification": "", "model": "ABC-123", "manufacturer_part_no": ""},
        ]
        self.assertEqual(exporter.exact_template_matches(template, sources), [sources[1]])

    def test_end_to_end_export_uses_existing_codes_and_covers_reviews(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_dir = root / "sources"
            evidence_dir = root / "evidence"
            source_dir.mkdir()
            evidence_dir.mkdir()

            a200 = source_dir / "A200量产BOM.xlsx"
            v700 = source_dir / "V700量产BOM.xlsx"
            self._write_source(a200, "M1809A_GM_A200_SUB_PCB_V3-BOM1 8SD05169C", [["芯片", "PART-A", 1, "U1"]])
            self._write_source(v700, "M1802A_GM_V700_TYPEC_V5-BOM1 6SD04989C", [["电阻", "10K", "", "R1-R3"], ["辅料", "胶", "", ""]])
            template = source_dir / "moban.xlsx"
            self._write_template(template)

            manifest_files = [self._snapshot(a200), self._snapshot(v700)]
            manifest = {"task": "TEST", "files": manifest_files, "manifest_sha256": exporter.canonical_sha(manifest_files)}
            manifest_path = evidence_dir / "source-manifest.json"
            manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")

            phone_like_hash = "3c258e" + "1" + "3" + "0" * 9 + "d3d"
            refs = [phone_like_hash, "ref-b", "ref-c"]
            profile_rows = [
                self._profile(refs[0], a200, 3, "芯片", "PART-A", "PART-A", "品牌A", "1", "U1"),
                self._profile(refs[1], v700, 3, "电阻", "10K", "R-10K", "品牌B", "", "R1-R3"),
                self._profile(refs[2], v700, 4, "辅料", "胶", "", "", "", ""),
            ]
            classification_rows = [
                self._classification(refs[0], a200, 3, "ELIGIBLE", "ELIGIBLE", "ELIGIBLE", "", "material:a", "CYD-RB_IC-000001", "RB_IC", "MPN"),
                self._classification(refs[1], v700, 3, "NEEDS_REVIEW", "ELIGIBLE", "NEEDS_REVIEW", "BOM_QUANTITY_NOT_VALID", "material:b", "CYD-RB_RES-000001", "RB_RES", "MPN"),
                self._classification(refs[2], v700, 4, "NEEDS_REVIEW", "NEEDS_REVIEW", "NEEDS_REVIEW", "CATEGORY_OR_COUNT_UNIT_NOT_DETERMINISTIC", "", "", "", ""),
            ]
            profile_path = evidence_dir / "profile.csv"
            classification_path = evidence_dir / "classification.csv"
            self._write_csv(profile_path, profile_rows)
            self._write_csv(classification_path, classification_rows)

            payload = {
                "manifest_sha256": manifest["manifest_sha256"],
                "materials": [
                    self._material("material:a", "CYD-RB_IC-000001", "RB_IC", "芯片", "PART-A", "品牌A", "PART-A", refs[0], a200, 3),
                    self._material("material:b", "CYD-RB_RES-000001", "RB_RES", "电阻", "10K", "品牌B", "R-10K", refs[1], v700, 3),
                ],
            }
            payload_path = evidence_dir / "payload.json"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            output = source_dir / "内部物料库.xlsx"
            report_path = evidence_dir / "report.json"
            before = {path.name: exporter.file_snapshot(path) for path in (a200, v700, template)}
            report = exporter.export_library(
                source_dir=source_dir,
                template_path=template,
                profile_csv=profile_path,
                classification_csv=classification_path,
                payload_json=payload_path,
                manifest_json=manifest_path,
                output=output,
                report_path=report_path,
                confirmation=exporter.CONFIRMATION,
            )
            after = {path.name: exporter.file_snapshot(path) for path in (a200, v700, template)}

            self.assertEqual(before, after)
            self.assertEqual(stat.S_IMODE(output.stat().st_mode), 0o600)
            self.assertEqual(report["counts"]["formal_materials"], 2)
            self.assertEqual(report["counts"]["material_review_source_rows"], 1)
            self.assertEqual(report["counts"]["classification_review_rows"], 2)
            self.assertEqual(report["counts"]["template_rows_strictly_mapped"], 1)
            self.assertEqual(report["validation"]["source_mapping_coverage"], "PASS")

            workbook = openpyxl.load_workbook(output, data_only=False, keep_links=False)
            try:
                self.assertEqual(workbook.sheetnames, ["内部物料库", "标准BOM明细", "待确认", "来源映射", "来源与说明"])
                details = workbook["标准BOM明细"]
                self.assertEqual(tuple(cell.value for cell in details[1]), exporter.TEMPLATE_HEADERS)
                self.assertEqual(details["K2"].value, '=IF(OR(G2="",J2=""),"",G2*J2)')
                self.assertTrue(all(row[3].value for row in details.iter_rows(min_row=2)))
                self.assertTrue(all(row[4].value for row in details.iter_rows(min_row=2)))
                library = workbook["内部物料库"]
                codes = [library.cell(row, 2).value for row in range(2, library.max_row + 1) if library.cell(row, 2).value]
                self.assertEqual(codes, ["CYD-RB_IC-000001", "CYD-RB_RES-000001"])
            finally:
                workbook.close()

    @staticmethod
    def _write_source(path, title, rows):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "BOM"
        sheet.append([title])
        sheet.append(["名称", "物料规格描述", "用量", "位号"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        workbook.close()
        path.chmod(0o444)

    @staticmethod
    def _write_template(path):
        workbook = openpyxl.Workbook()
        original = workbook.active
        original.title = "原BOM"
        sheet = workbook.create_sheet("Sheet1")
        sheet.append(list(exporter.TEMPLATE_HEADERS))
        sheet.append([1, "A200", "USB小板", "8SD05169C", "PART-A", "品牌A", 1, "", "供应商A", "", "", "", ""])
        workbook.save(path)
        workbook.close()
        path.chmod(0o444)

    @staticmethod
    def _snapshot(path):
        return exporter.file_snapshot(path)

    @staticmethod
    def _profile(source_ref, path, row, name, specification, mpn, maker, quantity, reference):
        return {
            "source_ref": source_ref,
            "file_sha256": exporter.sha256_path(path),
            "filename": path.name,
            "sheet": "BOM",
            "source_row": str(row),
            "classification": "NEEDS_REVIEW",
            "reason_codes": "",
            "mapping_rule_version": "test",
            "internal_stable_id": "",
            "internal_code": "",
            "material_name": name,
            "specification": specification,
            "model": "",
            "manufacturer_part_no": mpn,
            "manufacturer": maker,
            "package": "",
            "unit_raw": "",
            "unit": "",
            "quantity_raw": quantity,
            "quantity": quantity,
            "reference": reference,
            "substitute": "",
            "remark": "不应进入结果 " + "1" + "3" + "0" * 9,
            "version": "",
            "product": "",
        }

    @staticmethod
    def _classification(source_ref, path, row, overall, material, bom, reasons, stable_id, code, category, method):
        return {
            "source_ref": source_ref,
            "file_sha256": exporter.sha256_path(path),
            "sheet": "BOM",
            "source_row": str(row),
            "classification": overall,
            "material_classification": material,
            "bom_classification": bom,
            "reason_codes": reasons,
            "mapping_rule_version": "test",
            "internal_stable_id": stable_id,
            "internal_code": code,
            "category_code": category,
            "unit": "PCS" if code else "",
            "mapping_method": method,
        }

    @staticmethod
    def _material(stable, code, category, name, specification, maker, mpn, source_ref, path, row):
        return {
            "stable_key": stable,
            "internal_code": code,
            "category_code": category,
            "category_name": category,
            "identity_method": "MPN",
            "standard_name": name,
            "manufacturer": maker,
            "manufacturer_part_number": mpn,
            "specification": specification,
            "model": "",
            "package": "",
            "source_rows": [
                {
                    "source_ref": source_ref,
                    "file_sha256": exporter.sha256_path(path),
                    "sheet": "BOM",
                    "source_row": row,
                    "mapping_method": "MPN",
                }
            ],
        }

    @staticmethod
    def _write_csv(path, rows):
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(rows)


if __name__ == "__main__":
    unittest.main()
