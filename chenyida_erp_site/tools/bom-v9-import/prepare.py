#!/usr/bin/env python3
"""Prepare a root-only, explicit-field staging payload for the BOM V9 workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import stat
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RULE_VERSION = "bom-v9-explicit-fields-v1"
PAYLOAD_MARKER = "BOM_V9_EXPLICIT_STAGING_V1"
EXPECTED_SHEET = "ERP编码版"
REQUIRED_HEADERS = (
    "ERP物料编码",
    "物料大类",
    "物料名称",
    "规格型号",
    "封装",
    "参数",
    "电压",
    "材质",
    "精度",
    "使用次数",
    "项目来源",
    "原始描述",
)
UNIT_HEADERS = ("单位", "基本单位", "库存单位", "UOM", "Unit")
BOM_HEADERS = ("产品编码", "产品版本", "BOM版本", "BOM行号", "用量", "数量", "位号")
UNIT_ALIASES = {
    "PCS": "PCS",
    "PC": "PCS",
    "EA": "PCS",
    "件": "PCS",
    "个": "PCS",
}
ERP_CODE = re.compile(r"ERP-MAT-\d{5}\Z")


class PreparationError(RuntimeError):
    """Fail-closed workbook or output boundary violation."""


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def digest(value: Any) -> str:
    raw = value if isinstance(value, bytes) else canonical_json(value).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def file_sha256(path: Path) -> str:
    result = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            result.update(chunk)
    return result.hexdigest()


def file_stat(path: Path) -> dict[str, Any]:
    value = path.stat()
    return {
        "size": str(value.st_size),
        "mode": format(stat.S_IMODE(value.st_mode), "03o"),
        "uid": str(value.st_uid),
        "gid": str(value.st_gid),
        "mtime_ns": str(value.st_mtime_ns),
        "inode": str(value.st_ino),
    }


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalized(value: Any) -> str:
    return unicodedata.normalize("NFKC", cell_text(value)).strip()


def normalized_identity(value: Any) -> str:
    return re.sub(r"\s+", "", normalized(value)).upper()


def secure_output_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True, mode=0o700)
    path.chmod(0o700)
    value = path.stat()
    if value.st_uid != 0 or stat.S_IMODE(value.st_mode) != 0o700:
        raise PreparationError("输出目录必须为 root:root 0700")


def atomic_text(path: Path, content: str) -> None:
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    temporary.write_text(content, encoding="utf-8")
    temporary.chmod(0o600)
    temporary.replace(path)
    path.chmod(0o600)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    temporary.chmod(0o600)
    temporary.replace(path)
    path.chmod(0o600)


def explicit_unit(value: Any) -> str | None:
    key = normalized(value).upper()
    return UNIT_ALIASES.get(key)


def valid_usage_count(value: Any) -> bool:
    text = normalized(value)
    return bool(re.fullmatch(r"0|[1-9]\d*", text))


def prepare(workbook_path: Path, expected_sha256: str, output_dir: Path) -> dict[str, Any]:
    workbook_path = workbook_path.resolve(strict=True)
    if workbook_path.suffix.lower() != ".xlsx":
        raise PreparationError("只接受明确的 .xlsx 文件")
    if not re.fullmatch(r"[0-9a-f]{64}", expected_sha256):
        raise PreparationError("expected SHA-256 格式无效")

    secure_output_dir(output_dir)
    before = file_stat(workbook_path)
    actual_sha256 = file_sha256(workbook_path)
    if actual_sha256 != expected_sha256:
        raise PreparationError("原始文件 SHA-256 与授权值不一致")

    workbook = load_workbook(workbook_path, read_only=False, data_only=False, keep_links=True)
    try:
        if workbook.sheetnames != [EXPECTED_SHEET]:
            raise PreparationError("工作簿必须只有一个可见 ERP编码版 Sheet")
        worksheet = workbook[EXPECTED_SHEET]
        if worksheet.sheet_state != "visible":
            raise PreparationError("ERP编码版 Sheet 必须可见")
        if getattr(workbook, "_external_links", []):
            raise PreparationError("工作簿包含外部链接")

        formula_cells = [cell.coordinate for row in worksheet.iter_rows() for cell in row if cell.data_type == "f"]
        if formula_cells:
            raise PreparationError("工作簿包含公式，禁止猜测缓存值")

        header_values = [normalized(cell.value) for cell in worksheet[1]]
        while header_values and not header_values[-1]:
            header_values.pop()
        if len(header_values) != len(set(header_values)):
            raise PreparationError("表头包含重复字段")
        if tuple(header_values[: len(REQUIRED_HEADERS)]) != REQUIRED_HEADERS:
            raise PreparationError("V9 核心字段或顺序不符合受控契约")
        extras = header_values[len(REQUIRED_HEADERS) :]
        if any(header in BOM_HEADERS for header in extras):
            raise PreparationError("当前 staging 工具不解释产品/BOM 字段，必须使用独立显式契约")
        unknown_extras = [header for header in extras if header not in UNIT_HEADERS]
        if unknown_extras:
            raise PreparationError("存在未识别扩展字段")
        present_unit_headers = [header for header in UNIT_HEADERS if header in header_values]
        if len(present_unit_headers) > 1:
            raise PreparationError("存在多个单位字段，语义不唯一")
        unit_header = present_unit_headers[0] if present_unit_headers else None

        source_rows: list[dict[str, Any]] = []
        skipped_blank_rows = 0
        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2, max_col=len(header_values)), start=2):
            values = {header_values[index]: cell_text(cell.value) for index, cell in enumerate(cells)}
            if not any(normalized(value) for value in values.values()):
                skipped_blank_rows += 1
                continue
            source_rows.append({"source_row": row_number, "fields": values})
    finally:
        workbook.close()

    if not source_rows:
        raise PreparationError("工作簿没有数据行")

    code_counts = Counter(normalized(row["fields"].get("ERP物料编码")) for row in source_rows)
    identities: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for row in source_rows:
        fields = row["fields"]
        identity = tuple(normalized_identity(fields.get(header)) for header in REQUIRED_HEADERS[1:9])
        if any(identity):
            identities[identity].append(row["source_row"])
    duplicate_identity_rows = {
        row_number
        for row_numbers in identities.values()
        if len(row_numbers) > 1
        for row_number in row_numbers
    }

    staged_rows: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    reason_counts: Counter[str] = Counter()
    classification_counts: Counter[str] = Counter()
    for source in source_rows:
        row_number = int(source["source_row"])
        fields = source["fields"]
        code = normalized(fields.get("ERP物料编码"))
        reasons: list[str] = []
        if not code:
            reasons.append("ERP_CODE_MISSING")
        elif not ERP_CODE.fullmatch(code):
            reasons.append("ERP_CODE_INVALID")
        elif code_counts[code] > 1:
            reasons.append("ERP_CODE_DUPLICATE")
        if not normalized(fields.get("物料大类")):
            reasons.append("MATERIAL_CATEGORY_MISSING")
        if not normalized(fields.get("物料名称")):
            reasons.append("MATERIAL_NAME_MISSING")
        unit_value = fields.get(unit_header, "") if unit_header else ""
        unit_code = explicit_unit(unit_value)
        if not unit_header or not normalized(unit_value):
            reasons.append("EXPLICIT_UNIT_MISSING")
        elif not unit_code:
            reasons.append("EXPLICIT_UNIT_UNSUPPORTED")
        if not valid_usage_count(fields.get("使用次数")):
            reasons.append("USAGE_COUNT_INVALID")
        if not normalized(fields.get("项目来源")) or not normalized(fields.get("原始描述")):
            reasons.append("SOURCE_TRACE_MISSING")
        if row_number in duplicate_identity_rows:
            reasons.append("EXACT_IDENTITY_DUPLICATE_REVIEW_REQUIRED")

        reasons = sorted(set(reasons))
        classification = "NEEDS_REVIEW" if reasons else "ELIGIBLE"
        source_ref = f"xlsx:{actual_sha256}:{EXPECTED_SHEET}:{row_number}"
        row_snapshot = {header: fields.get(header, "") for header in header_values}
        row_digest = digest(row_snapshot)
        staged = {
            "source_ref": source_ref,
            "source_row": row_number,
            "row_digest": row_digest,
            "erp_material_code": code,
            "classification": classification,
            "reason_codes": reasons,
            "unit_code": unit_code or "",
            "usage_count": normalized(fields.get("使用次数")),
            "row_snapshot": row_snapshot,
        }
        staged_rows.append(staged)
        classification_counts[classification] += 1
        reason_counts.update(reasons)
        if classification == "NEEDS_REVIEW":
            review_rows.append(
                {
                    "source_ref": source_ref,
                    "source_row": row_number,
                    "erp_material_code": code,
                    "material_category": normalized(fields.get("物料大类")),
                    "material_name": normalized(fields.get("物料名称")),
                    "specification_model": normalized(fields.get("规格型号")),
                    "classification": classification,
                    "reason_codes": "|".join(reasons),
                }
            )

    codes = [normalized(row["fields"].get("ERP物料编码")) for row in source_rows]
    valid_code_numbers = sorted(
        int(code.rsplit("-", 1)[1]) for code in codes if ERP_CODE.fullmatch(code)
    )
    missing_sequence_count = 0
    if valid_code_numbers:
        missing_sequence_count = len(
            set(range(valid_code_numbers[0], valid_code_numbers[-1] + 1))
            - set(valid_code_numbers)
        )
    explicit_unit_rows = sum(bool(row["unit_code"]) for row in staged_rows)
    source_trace_rows = sum(
        bool(normalized(row["fields"].get("项目来源")))
        and bool(normalized(row["fields"].get("原始描述")))
        for row in source_rows
    )
    usage_total = sum(
        int(normalized(row["fields"].get("使用次数")))
        for row in source_rows
        if valid_usage_count(row["fields"].get("使用次数"))
    )
    workbook_issues: list[str] = []
    if not unit_header:
        workbook_issues.append("EXPLICIT_UNIT_COLUMN_MISSING")
    if not all(header in header_values for header in BOM_HEADERS):
        workbook_issues.append("BOM_STRUCTURE_NOT_PRESENT")

    summary = {
        "source_rows": len(source_rows),
        "skipped_blank_rows": skipped_blank_rows,
        "classification_counts": dict(sorted(classification_counts.items())),
        "reason_counts": dict(sorted(reason_counts.items())),
        "erp_codes": {
            "present": sum(bool(code) for code in codes),
            "unique": len(set(code for code in codes if code)),
            "invalid": sum(bool(code) and not ERP_CODE.fullmatch(code) for code in codes),
            "duplicate_rows": sum(count for count in code_counts.values() if count > 1),
            "sequence_min": valid_code_numbers[0] if valid_code_numbers else None,
            "sequence_max": valid_code_numbers[-1] if valid_code_numbers else None,
            "missing_sequence_count": missing_sequence_count,
        },
        "exact_identity_duplicate_groups": sum(1 for rows in identities.values() if len(rows) > 1),
        "explicit_unit_rows": explicit_unit_rows,
        "missing_or_unsupported_unit_rows": len(staged_rows) - explicit_unit_rows,
        "valid_usage_count_rows": sum(valid_usage_count(row["fields"].get("使用次数")) for row in source_rows),
        "usage_count_total_trace_only": usage_total,
        "source_trace_rows": source_trace_rows,
        "materials_ready": classification_counts.get("ELIGIBLE", 0),
        "products": 0,
        "product_versions": 0,
        "boms": 0,
        "bom_versions": 0,
        "bom_lines": 0,
        "workbook_issues": workbook_issues,
    }
    manifest = {
        "schema_version": 1,
        "source_file": {
            "filename": workbook_path.name,
            "sha256": actual_sha256,
            **before,
        },
        "sheet": EXPECTED_SHEET,
        "headers": header_values,
    }
    manifest["manifest_sha256"] = digest(manifest)
    payload = {
        "schema_version": 1,
        "marker": PAYLOAD_MARKER,
        "rule_version": RULE_VERSION,
        "manifest": manifest,
        "summary": summary,
        "rows": staged_rows,
    }
    payload["payload_digest"] = digest(payload)

    after = file_stat(workbook_path)
    after_sha256 = file_sha256(workbook_path)
    if before != after or actual_sha256 != after_sha256:
        raise PreparationError("原始工作簿在解析期间发生变化")

    atomic_text(output_dir / "source-manifest.json", canonical_json(manifest) + "\n")
    atomic_text(output_dir / "staging-payload.json", canonical_json(payload) + "\n")
    atomic_text(output_dir / "staging-summary.json", canonical_json(summary) + "\n")
    write_csv(
        output_dir / "needs-review.csv",
        [
            "source_ref",
            "source_row",
            "erp_material_code",
            "material_category",
            "material_name",
            "specification_model",
            "classification",
            "reason_codes",
        ],
        review_rows,
    )
    return summary


def cli() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--expected-sha256", required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    arguments = parser.parse_args()
    try:
        summary = prepare(arguments.workbook, arguments.expected_sha256, arguments.output_dir)
    except PreparationError as error:
        print(canonical_json({"ok": False, "code": "BOM_V9_PREPARATION_BLOCKED", "message": str(error)}))
        return 1
    print(canonical_json({"ok": True, "summary": summary}))
    return 0


if __name__ == "__main__":
    raise SystemExit(cli())
