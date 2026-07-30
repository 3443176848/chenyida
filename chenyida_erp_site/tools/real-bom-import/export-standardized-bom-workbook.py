#!/usr/bin/env python3
"""Standardize every guarded BOM source using the moban raw-to-clean pattern.

The exporter is deliberately offline.  It reads immutable spreadsheet sources,
learns and validates the row-grouping contract demonstrated by ``moban.xlsx``
(``原BOM`` -> ``Sheet1``), writes one canonical 13-column sheet per logical
source, and builds one non-deduplicated consolidated sheet.  It never connects
to a database and never assigns or changes a formal ERP material code.
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import posixpath
import re
import stat
import tempfile
import unicodedata
from collections import Counter
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TASK = "SELFHOST-LANDING-TASK07"
CONFIRMATION = "OFFLINE_STANDARDIZED_BOM_WORKBOOK_EXPORT"
TEMPLATE_HEADERS = (
    "序号",
    "项目号",
    "板子类型",
    "内部型号",
    "物料规格描述",
    "品牌",
    "用量",
    "替代料",
    "供应商",
    "订单数量",
    "需求数量",
    "购买数量",
    "库存数",
)
SOURCE_SHEET_ORDER = (
    "1928C标准",
    "A118标准",
    "A200量产BOM标准",
    "A200物料清单标准",
    "A200量产注意事项",
    "G20-G15G标准",
    "J587标准",
    "V700标准",
)
EXPECTED_SOURCE_FILES = {
    "1928C量产BOM.xlsx",
    "A118量产BOM.xlsx",
    "A200量产BOM.xlsx",
    "A200量产注意事项.xls",
    "A200量产物料清单.xlsx",
    "G20-G15G项目量产BOM.xlsx",
    "J587_SUBA2_V01-20260703.xlsx",
    "V700量产BOM.xlsx",
}

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
PHONE_LIKE_RE = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")
SENSITIVE_RE = re.compile(
    r"(?:postgres(?:ql)?://|database_url|password\s*[=:]|api[_-]?key\s*[=:]|bearer\s+[a-z0-9._-]+)",
    re.I,
)
MODEL_RE = re.compile(r"(?<![A-Z0-9])(\d(?:SD|SH|PF|P)\d{5,6}[A-Z]?)(?![A-Z0-9])", re.I)
J587_MODEL_RE = re.compile(r"(J587[_-]SUBA\d+[_-]V\d+)", re.I)
REF_RE = re.compile(r"(?<![A-Z0-9])([A-Z]{1,5}\d+(?:[-_]\d+)?)(?![A-Z0-9])", re.I)


@dataclass(frozen=True)
class Context:
    project: str
    board_type: str
    internal_model: str
    title: str


@dataclass
class RawLine:
    filename: str
    sheet: str
    row_no: int
    context: Context
    sequence: str = ""
    status: str = ""
    source_code: str = ""
    name: str = ""
    specification: str = ""
    manufacturer_part_no: str = ""
    brand: str = ""
    quantity_raw: Any = None
    reference: str = ""
    remark: str = ""
    substitute: str = ""


@dataclass
class RawGroup:
    primary: RawLine
    alternatives: list[RawLine] = field(default_factory=list)
    quantity: int | float | None = None
    quantity_source: str = ""


@dataclass
class StandardRow:
    project: str
    board_type: str
    internal_model: str
    specification: str
    brand: str
    quantity: int | float | None
    substitute: str
    supplier: str
    order_quantity: int | float | None = None
    inventory: int | float | None = None
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    alternative_rows: tuple[int, ...] = ()
    rule: str = ""
    status: str = "已按模板规则整理"


@dataclass(frozen=True)
class TemplateRow:
    source_row: int
    project: str
    board_type: str
    internal_model: str
    specification: str
    brand: str
    quantity: int | float | None
    substitute: str
    supplier: str
    order_quantity: int | float | None
    inventory: int | float | None


@dataclass
class ParseResult:
    rows: list[StandardRow]
    anomalies: list[dict[str, Any]]
    stats: dict[str, Any]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return unicodedata.normalize("NFC", str(value)).strip()


def normalized_header(value: Any) -> str:
    return re.sub(r"[\s\-_/\\:：()（）\[\]【】.]+", "", text(value).lower())


def strict_key(value: Any) -> str:
    result = unicodedata.normalize("NFKC", text(value)).upper()
    result = result.translate(str.maketrans({"，": ",", "；": ";", "：": ":", "。": "."}))
    return re.sub(r"[^0-9A-Z\u4e00-\u9fff.+/%±]+", "", result)


def safe_excel_text(value: Any, *, limit: int = 30000) -> str:
    result = CONTROL_RE.sub("", text(value))
    result = PHONE_LIKE_RE.sub("[已移除可能的联系电话]", result)
    if result.lstrip().startswith(("=", "+", "-", "@")):
        result = "'" + result
    return result[:limit]


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_sha(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def file_snapshot(path: Path) -> dict[str, Any]:
    info = path.lstat()
    if not stat.S_ISREG(info.st_mode) or path.is_symlink():
        raise ValueError(f"INPUT_NOT_REGULAR:{path.name}")
    return {
        "filename": path.name,
        "sha256": sha256_path(path),
        "size_bytes": info.st_size,
        "inode": info.st_ino,
        "mode": oct(stat.S_IMODE(info.st_mode)),
        "uid": info.st_uid,
        "gid": info.st_gid,
        "mtime_ns": info.st_mtime_ns,
    }


def verify_manifest(source_dir: Path, manifest: dict[str, Any]) -> dict[str, dict[str, Any]]:
    files = manifest.get("files")
    if not isinstance(files, list) or not files:
        raise ValueError("SOURCE_MANIFEST_EMPTY")
    if canonical_sha(files) != manifest.get("manifest_sha256"):
        raise ValueError("SOURCE_MANIFEST_DIGEST_MISMATCH")
    filenames = {text(item.get("filename")) for item in files}
    if filenames != EXPECTED_SOURCE_FILES:
        raise ValueError("SOURCE_MANIFEST_FILE_SET_MISMATCH")
    snapshots: dict[str, dict[str, Any]] = {}
    for expected in files:
        filename = text(expected.get("filename"))
        path = source_dir / filename
        actual = file_snapshot(path)
        for key in ("filename", "sha256", "size_bytes", "inode", "mode", "uid", "gid", "mtime_ns"):
            if str(actual[key]) != str(expected.get(key)):
                raise ValueError(f"SOURCE_FILE_DRIFT:{filename}:{key}")
        snapshots[filename] = actual
    return snapshots


def excel_number(value: Any, *, positive: bool = False) -> int | float | None:
    raw = text(value).replace(",", "")
    if not raw or raw.startswith("="):
        return None
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return None
    if not number.is_finite() or number < 0 or (positive and number <= 0):
        return None
    if -number.as_tuple().exponent > 6:
        return None
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def column_number(reference: str) -> int:
    match = re.match(r"([A-Z]+)", reference)
    if not match:
        return 0
    result = 0
    for char in match.group(1):
        result = result * 26 + ord(char) - 64
    return result


def _xlsx_scalar(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    value_node = cell.find(f"{{{NS_MAIN}}}v")
    if cell_type == "s" and value_node is not None and value_node.text is not None:
        return shared[int(value_node.text)]
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iter(f"{{{NS_MAIN}}}t"))
    if cell_type == "b" and value_node is not None:
        return value_node.text == "1"
    if value_node is None or value_node.text is None:
        return ""
    raw = value_node.text
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return raw
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def sparse_xlsx_cells(path: Path, *, max_column: int = 16) -> dict[str, dict[int, dict[int, Any]]]:
    result: dict[str, dict[int, dict[int, Any]]] = {}
    with ZipFile(path) as archive:
        names = set(archive.namelist())
        if any(name.endswith("vbaProject.bin") or name.startswith("xl/externalLinks/") for name in names):
            raise ValueError(f"UNSAFE_XLSX_FEATURE:{path.name}")
        if any(name.startswith("/") or ".." in Path(name).parts for name in names):
            raise ValueError(f"UNSAFE_XLSX_PATH:{path.name}")
        shared: list[str] = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared = ["".join(node.text or "" for node in item.iter(f"{{{NS_MAIN}}}t")) for item in root]
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {node.attrib["Id"]: node.attrib["Target"] for node in relationships}
        sheets = workbook.find(f"{{{NS_MAIN}}}sheets")
        if sheets is None:
            return result
        for sheet in sheets:
            relationship_id = sheet.attrib[f"{{{NS_REL}}}id"]
            target = targets[relationship_id].lstrip("/")
            xml_path = target if target.startswith("xl/") else posixpath.normpath(f"xl/{target}")
            rows: dict[int, dict[int, Any]] = {}
            for _event, element in ET.iterparse(io.BytesIO(archive.read(xml_path)), events=("end",)):
                if element.tag != f"{{{NS_MAIN}}}row":
                    continue
                row_values: dict[int, Any] = {}
                for cell in element.findall(f"{{{NS_MAIN}}}c"):
                    column = column_number(cell.attrib.get("r", ""))
                    if not 0 < column <= max_column:
                        continue
                    value = _xlsx_scalar(cell, shared)
                    if text(value):
                        row_values[column] = value
                if row_values:
                    rows[int(element.attrib["r"])] = row_values
                element.clear()
            result[sheet.attrib["name"]] = rows
    return result


HEADER_ALIASES: dict[str, set[str]] = {
    "sequence": {"序号", "项次", "item", "no", "编号"},
    "status": {"状态"},
    "source_code": {"物料编码", "内部物料编码", "hc_code", "hccode", "物料编号"},
    "name": {"名称", "物料名称", "品名"},
    "specification": {"物料规格描述", "物料描述", "规格描述", "规格", "description", "描述", "规格型号", "物料名称及描述"},
    "manufacturer_part_no": {"物料型号", "型号", "厂商物料编码", "vendorcode", "供应商料号"},
    "brand": {"生产厂商", "生产厂家", "制造商", "厂家", "品牌", "vendor"},
    "quantity": {"单机用量", "用量", "数量", "qty", "quantity", "普通用量"},
    "reference": {"位号", "位置", "reference", "ref", "普通用量位置", "用量位置"},
    "remark": {"备注"},
    "substitute": {"替代料"},
}


def header_field(value: Any) -> str | None:
    token = normalized_header(value)
    if not token:
        return None
    for field_name, aliases in HEADER_ALIASES.items():
        if token in {normalized_header(alias) for alias in aliases}:
            return field_name
    return None


def header_mapping(cells: dict[int, Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column, value in sorted(cells.items()):
        field_name = header_field(value)
        if field_name and field_name not in mapping:
            mapping[field_name] = column
    return mapping


def is_header_mapping(mapping: dict[str, int]) -> bool:
    if len(mapping) >= 3 and ("specification" in mapping or "name" in mapping):
        return True
    return "specification" in mapping and "quantity" in mapping


def row_join(cells: dict[int, Any]) -> str:
    return " | ".join(text(value) for _column, value in sorted(cells.items()) if text(value))


def is_section_title(cells: dict[int, Any]) -> bool:
    values = [text(value) for value in cells.values() if text(value)]
    if not values or len(values) > 4:
        return False
    combined = " ".join(values)
    upper = unicodedata.normalize("NFKC", combined).upper()
    if MODEL_RE.search(upper) or J587_MODEL_RE.search(upper):
        return True
    if "BOM" in upper and not upper.startswith("BOM功能描述"):
        return True
    if re.search(r"\b\d{4,6}[A-Z]\b", upper) and any(marker in upper for marker in ("A118", "USB", "FPC", "小板")):
        return True
    return False


def model_from_title(title: str, *, project: str, fallback: str = "") -> str:
    upper = unicodedata.normalize("NFKC", title).upper()
    matches = MODEL_RE.findall(upper)
    if matches:
        return matches[-1]
    j587 = J587_MODEL_RE.search(upper)
    if j587:
        return j587.group(1).replace("-", "_")
    if project == "A118" and "A118" in upper:
        short = re.search(r"\b(\d{4,6}[A-Z])\b", upper)
        if short:
            return short.group(1)
    return fallback


def board_type_from_title(title: str, sheet: str, project: str) -> str:
    upper = unicodedata.normalize("NFKC", f"{title} {sheet}").upper()
    patterns = (
        (("PSSENSOR", "光感"), "光感小板"),
        (("SUBLCM",), "屏小板"),
        (("屏排线",), "屏排线"),
        (("主FPC",), "主FPC"),
        (("充电FPC",), "充电FPC"),
        (("SIM-FPC", "SIM FPC"), "SIM FPC"),
        (("IRLED",), "IRLED FPC"),
        (("自定义键",), "自定义键"),
        (("侧键",), "侧键"),
        (("TYPE-C耳机",), "TYPE-C耳机小板"),
        (("TYPE-C", "TYPEC"), "TYPE-C小板"),
        (("USB",), "USB小板"),
        (("DMR&RTK", "2WDMR&RTK"), "DMR/RTK小板"),
        (("DMR",), "DMR/LTE小板"),
        (("MOTOR",), "马达小板"),
        (("WIFI",), "WiFi小板"),
        (("N102",), "N102小板"),
        (("座充", "-CH-"), "座充指示小板"),
        (("5G小板",), "5G小板"),
        (("CX-78",), "CX-78小板"),
        (("单闪",), "单闪小板"),
        (("双闪",), "双闪小板"),
        (("闪光灯",), "闪光灯"),
        (("ANT1",), "ANT1天线小板"),
        (("ANT3",), "ANT3天线小板"),
        (("K509-SUB", "_SUB_", "-SUB-"), "SUB小板"),
        (("ANT", "天线"), "天线小板"),
        (("SUB",), "SUB小板"),
    )
    for markers, label in patterns:
        if any(marker in upper for marker in markers):
            return label
    if project == "J587":
        return "TYPE-C耳机小板"
    return ""


def context_from_title(project: str, sheet: str, title: str, previous: Context, fallback_model: str) -> Context:
    model = model_from_title(title, project=project, fallback=fallback_model)
    board_type = board_type_from_title(title, sheet, project)
    if project == "A200" and model:
        board_type = {
            "8SD05169C": "USB小板",
            "4P05171B": "USB FPC",
            "3PF05170C": "主FPC",
            "2SH05177A": "光感小板",
            "2P05175A": "自定义键",
            "2P05176A": "侧键",
            "2P05173D": "屏排线",
            "2P05174B": "闪光灯",
        }.get(model, board_type)
    if not model and previous.internal_model:
        model = previous.internal_model
    if not board_type and not model:
        board_type = previous.board_type
    return Context(project=project, board_type=board_type, internal_model=model, title=title)


def looks_like_reference(value: Any) -> bool:
    source = unicodedata.normalize("NFKC", text(value)).upper()
    return bool(REF_RE.search(source))


def reference_quantity(value: Any) -> int | None:
    source = unicodedata.normalize("NFKC", text(value)).upper()
    if not source:
        return None
    tokens = REF_RE.findall(source)
    if not tokens:
        return None
    unique = list(dict.fromkeys(tokens))
    return len(unique)


def meaningful_material(line: RawLine) -> bool:
    combined = " ".join((line.name, line.specification, line.manufacturer_part_no, line.source_code)).strip()
    if not combined:
        return False
    token = normalized_header(combined)
    if token in {"合计", "总计", "小计", "物料规格描述", "物料名称及描述"}:
        return False
    return True


def is_board_base(line: RawLine) -> bool:
    if looks_like_reference(line.reference):
        return False
    combined = unicodedata.normalize("NFKC", f"{line.name} {line.specification}").upper()
    return (
        bool(re.search(r"(?:^|[^A-Z])PCB(?:A)?(?:[^A-Z]|$)", combined))
        or "空板" in combined
        or "小板组件" in combined
        or "主板/板贴" in combined
    )


def infer_a118_mapping(cells: dict[int, Any], current: dict[str, int]) -> dict[str, int]:
    quantity = excel_number(cells.get(current.get("quantity", -1)), positive=True)
    if quantity is not None:
        return current
    if excel_number(cells.get(6), positive=True) is not None and (text(cells.get(2)) or text(cells.get(3))):
        return {
            "sequence": 1,
            "name": 2,
            "specification": 3,
            "manufacturer_part_no": 4,
            "brand": 5,
            "quantity": 6,
            "reference": 7,
            "remark": 8,
            "substitute": 9,
        }
    if excel_number(cells.get(5), positive=True) is not None and text(cells.get(3)):
        return {
            "sequence": 1,
            "source_code": 2,
            "specification": 3,
            "manufacturer_part_no": 4,
            "quantity": 5,
            "reference": 6,
            "remark": 7,
            "substitute": 8,
        }
    return current


def brand_from_specification(specification: str, manufacturer_part_no: str) -> str:
    """Recover an explicit trailing manufacturer when the source has no brand column."""

    source = unicodedata.normalize("NFKC", specification)
    parts = [part.strip() for part in re.split(r"[,，;；]", source) if part.strip()]
    if len(parts) < 2:
        return ""
    candidate = parts[-1]
    if strict_key(candidate) == strict_key(manufacturer_part_no):
        return ""
    mpn_precedes_candidate = bool(
        manufacturer_part_no
        and any(strict_key(manufacturer_part_no) in strict_key(part) for part in parts[:-1])
    )
    upper = candidate.upper()
    if len(candidate) > 32 or re.search(r"\d+(?:\.\d+)?\s*(?:MM|V|A|W|PF|NF|UF|NH|OHM|PIN)\b", upper):
        return ""
    if (
        not mpn_precedes_candidate
        and re.fullmatch(r"[A-Z0-9_.+/-]{8,}", upper)
        and not re.search(r"(?:SEMI|TECH|LRC|ST|TDK|JAE|LCN|KOA)$", upper)
    ):
        return ""
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", candidate):
        return ""
    return candidate


def line_from_cells(
    *,
    filename: str,
    sheet: str,
    row_no: int,
    cells: dict[int, Any],
    mapping: dict[str, int],
    context: Context,
) -> RawLine:
    def get(field_name: str) -> Any:
        column = mapping.get(field_name)
        return cells.get(column) if column else None

    specification = text(get("specification"))
    manufacturer_part_no = text(get("manufacturer_part_no"))
    brand = text(get("brand")) or brand_from_specification(specification, manufacturer_part_no)
    return RawLine(
        filename=filename,
        sheet=sheet,
        row_no=row_no,
        context=context,
        sequence=text(get("sequence")),
        status=text(get("status")),
        source_code=text(get("source_code")),
        name=text(get("name")),
        specification=specification,
        manufacturer_part_no=manufacturer_part_no,
        brand=brand,
        quantity_raw=get("quantity"),
        reference=text(get("reference")),
        remark=text(get("remark")),
        substitute=text(get("substitute")),
    )


def is_primary(line: RawLine) -> bool:
    status = normalized_header(line.status)
    if status in {"主料", "main", "primary"}:
        return True
    if excel_number(line.quantity_raw, positive=True) is not None:
        return True
    return looks_like_reference(line.reference)


def is_alternative(line: RawLine) -> bool:
    status = normalized_header(line.status)
    if status in {"替代料", "替代", "alternative", "alternate"}:
        return True
    return bool(line.manufacturer_part_no or (line.source_code and line.specification))


def finalize_group(group: RawGroup | None, groups: list[RawGroup], anomalies: list[dict[str, Any]]) -> None:
    if group is None:
        return
    quantity = excel_number(group.primary.quantity_raw, positive=True)
    source = "原表用量"
    if quantity is None:
        quantity = reference_quantity(group.primary.reference)
        source = "位号计数" if quantity is not None else "未知"
    group.quantity = quantity
    group.quantity_source = source
    if quantity is None:
        anomalies.append(
            {
                "filename": group.primary.filename,
                "sheet": group.primary.sheet,
                "row": group.primary.row_no,
                "project": group.primary.context.project,
                "internal_model": group.primary.context.internal_model,
                "kind": "用量待确认",
                "detail": "原表未提供可验证的正数用量或可数位号；输出用量留空。",
            }
        )
    groups.append(group)


def parse_classic_groups(
    path: Path,
    *,
    project: str,
    sheet_name: str,
    fallback_model: str = "",
    special_j587: bool = False,
) -> tuple[list[RawGroup], list[dict[str, Any]], dict[str, Any]]:
    workbook = sparse_xlsx_cells(path, max_column=16)
    if sheet_name not in workbook:
        raise ValueError(f"SOURCE_SHEET_MISSING:{path.name}:{sheet_name}")
    source_rows = workbook[sheet_name]
    context = Context(project=project, board_type=board_type_from_title("", sheet_name, project), internal_model=fallback_model, title="")
    active_mapping: dict[str, int] = {}
    groups: list[RawGroup] = []
    anomalies: list[dict[str, Any]] = []
    current_group: RawGroup | None = None
    ignored_unassigned = 0
    ignored_unassigned_rows: list[int] = []
    board_base_excluded = 0
    for row_no, cells in sorted(source_rows.items()):
        mapping_candidate = header_mapping(cells)
        if is_header_mapping(mapping_candidate):
            finalize_group(current_group, groups, anomalies)
            current_group = None
            active_mapping = mapping_candidate
            if special_j587:
                active_mapping.setdefault("quantity", 9)
            continue
        if is_section_title(cells):
            finalize_group(current_group, groups, anomalies)
            current_group = None
            title = row_join(cells)
            context = context_from_title(project, sheet_name, title, context, fallback_model)
            continue
        if not active_mapping:
            continue
        row_mapping = infer_a118_mapping(cells, active_mapping) if project == "A118" else active_mapping
        line = line_from_cells(
            filename=path.name,
            sheet=sheet_name,
            row_no=row_no,
            cells=cells,
            mapping=row_mapping,
            context=context,
        )
        if not meaningful_material(line):
            continue
        if is_board_base(line):
            board_base_excluded += 1
            anomalies.append(
                {
                    "filename": path.name,
                    "sheet": sheet_name,
                    "row": row_no,
                    "project": project,
                    "internal_model": context.internal_model,
                    "kind": "板件本体未计入",
                    "detail": "按模板规则，PCB/PCBA/空板本体不作为采购物料明细写入汇总。",
                }
            )
            continue
        if is_primary(line):
            finalize_group(current_group, groups, anomalies)
            current_group = None
            current_group = RawGroup(primary=line)
            continue
        if is_alternative(line) and current_group is not None:
            current_group.alternatives.append(line)
            continue
        ignored_unassigned += 1
        ignored_unassigned_rows.append(row_no)
    finalize_group(current_group, groups, anomalies)
    stats = {
        "source_nonempty_rows": len(source_rows),
        "groups": len(groups),
        "alternative_rows": sum(len(group.alternatives) for group in groups),
        "board_base_excluded": board_base_excluded,
        "unassigned_nonprimary_rows": ignored_unassigned,
        "unassigned_nonprimary_row_numbers": ignored_unassigned_rows,
    }
    return groups, anomalies, stats


def material_category(name: str, specification: str) -> str:
    upper = unicodedata.normalize("NFKC", f"{name} {specification}").upper()
    if any(token in upper for token in ("热敏", "THERMISTOR")):
        return "thermistor"
    if any(token in upper for token in ("电阻", "RESISTOR", "RES-CHIP", "RES FILM", "RES,")):
        return "resistor"
    if any(token in upper for token in ("电容", "CAPACITOR", "CAP CER", "CAP,")):
        return "capacitor"
    if any(token in upper for token in ("磁珠", "FERRITE", "BEAD")):
        return "bead"
    if any(token in upper for token in ("电感", "INDUCTOR", "IND CER", "IND,")):
        return "inductor"
    if any(token in upper for token in ("连接器", "CONNECTOR", "BTB", "TYPE-C", "TYPE C", "USB", "卡座", "耳机座", "同轴座")):
        return "connector"
    if any(token in upper for token in ("TVS", "ESD", "浪涌管", "二极管", "DIODE")):
        return "tvs"
    if any(token in upper for token in ("IC", "SENSOR", "MIC", "传感", "硅麦", "模块")):
        return "active"
    if any(token in upper for token in ("LED", "灯", "闪")):
        return "led"
    if any(token in upper for token in ("SWITCH", "按键", "开关")):
        return "switch"
    return "other"


def remove_token(source: str, token: str) -> str:
    if not token:
        return source
    return re.sub(re.escape(token), "", source, flags=re.I)


def tidy_spec(value: str) -> str:
    result = unicodedata.normalize("NFKC", text(value))
    result = result.translate(str.maketrans({"，": ",", "；": ";", "：": ":"}))
    result = re.sub(r"\s+", " ", result)
    result = re.sub(r"\s*[,;]\s*", ",", result)
    result = re.sub(r",{2,}", ",", result)
    return result.strip(" ,;:_-")


def standardized_specification(line: RawLine) -> tuple[str, str]:
    category = material_category(line.name, line.specification)
    source = line.specification or line.name or line.manufacturer_part_no or line.source_code
    source = remove_token(source, line.brand)
    if category in {"resistor", "capacitor", "inductor", "thermistor", "bead"}:
        source = remove_token(source, line.manufacturer_part_no)
        source = source.replace("_", ",")
        replacements = (
            (r"^RES(?:-CHIP|\s+FILM|ISTOR)?\s*[,;-]?", "电阻,"),
            (r"^CAP(?:ACITOR|\s+CER)?\s*[,;-]?", "电容,"),
            (r"^IND(?:UCTOR|\s+CER)?\s*[,;-]?", "电感,"),
        )
        source = unicodedata.normalize("NFKC", source)
        for pattern, replacement in replacements:
            source = re.sub(pattern, replacement, source, flags=re.I)
        prefixes = {
            "resistor": "电阻",
            "capacitor": "电容",
            "inductor": "电感",
            "thermistor": "热敏电阻",
            "bead": "磁珠",
        }
        if strict_key(prefixes[category]) not in strict_key(source):
            source = f"{prefixes[category]},{source}"
        return tidy_spec(source), "无源器件保留通用电气规格，料号/品牌分离"
    if category in {"active", "tvs", "led"} and line.manufacturer_part_no:
        return text(line.manufacturer_part_no), "IC/传感/保护/灯类以关键料号作为规格"
    source = tidy_spec(source)
    if line.manufacturer_part_no and strict_key(line.manufacturer_part_no) not in strict_key(source):
        prefix = tidy_spec(line.name) or {
            "connector": "连接器",
            "switch": "开关",
        }.get(category, "")
        source = tidy_spec(f"{prefix},{line.manufacturer_part_no}")
    return source, "结构/连接器类保留名称、关键规格和料号"


def substitute_text(group: RawGroup) -> str:
    values: list[str] = []

    def append(value: str) -> None:
        value = re.sub(r"\s+", " ", unicodedata.normalize("NFKC", text(value))).strip(" ,;:")
        if not value:
            return
        key = strict_key(value)
        if key and all(strict_key(existing) != key for existing in values):
            values.append(value)

    append(group.primary.substitute)
    for alternative in group.alternatives:
        identity = alternative.manufacturer_part_no or alternative.source_code or alternative.specification
        if identity and alternative.brand:
            identity = f"{identity}（{alternative.brand}）"
        append(identity)
    return "；".join(values)[:3000]


def standard_rows_from_groups(groups: Iterable[RawGroup]) -> list[StandardRow]:
    rows: list[StandardRow] = []
    for group in groups:
        line = group.primary
        specification_line = line
        if material_category(line.name, line.specification) == "other":
            replacement = next(
                (
                    alternative
                    for alternative in group.alternatives
                    if material_category(alternative.name, alternative.specification) != "other"
                ),
                None,
            )
            if replacement is not None:
                specification_line = replacement
        specification, rule = standardized_specification(specification_line)
        if specification_line is not line:
            rule += "；主料描述不足，通用规格取同组明确备选料"
        rows.append(
            StandardRow(
                project=line.context.project,
                board_type=line.context.board_type,
                internal_model=line.context.internal_model,
                specification=specification,
                brand=line.brand,
                quantity=group.quantity,
                substitute=substitute_text(group),
                supplier="",
                source_file=line.filename,
                source_sheet=line.sheet,
                source_row=line.row_no,
                alternative_rows=tuple(item.row_no for item in group.alternatives),
                rule=f"{rule}；用量来源：{group.quantity_source}",
            )
        )
    return rows


def raw_group_signature(group: RawGroup) -> tuple[Any, ...]:
    def line_signature(line: RawLine) -> tuple[str, ...]:
        return (
            strict_key(line.source_code),
            strict_key(line.name),
            strict_key(line.specification),
            strict_key(line.manufacturer_part_no),
            strict_key(line.brand),
            strict_key(line.reference),
            strict_key(line.substitute),
        )

    return (
        line_signature(group.primary),
        group.quantity,
        tuple(line_signature(line) for line in group.alternatives),
    )


def suppress_identical_repeated_sections(
    groups: list[RawGroup],
) -> tuple[list[RawGroup], list[dict[str, Any]], dict[str, int]]:
    if not groups:
        return [], [], {"duplicate_sections_suppressed": 0, "duplicate_groups_suppressed": 0}
    blocks: list[tuple[tuple[str, str, str], list[RawGroup]]] = []
    for group in groups:
        context = group.primary.context
        key = (context.project, context.internal_model, strict_key(context.title))
        if not blocks or blocks[-1][0] != key:
            blocks.append((key, [group]))
        else:
            blocks[-1][1].append(group)
    seen: dict[tuple[str, str, str], tuple[tuple[Any, ...], int, int]] = {}
    output: list[RawGroup] = []
    anomalies: list[dict[str, Any]] = []
    duplicate_sections = 0
    duplicate_groups = 0
    for key, block in blocks:
        signature = tuple(raw_group_signature(group) for group in block)
        start_row = block[0].primary.row_no
        end_row = block[-1].primary.row_no
        prior = seen.get(key)
        if prior and prior[0] == signature:
            duplicate_sections += 1
            duplicate_groups += len(block)
            first_start, first_end = prior[1], prior[2]
            primary = block[0].primary
            anomalies.append(
                {
                    "filename": primary.filename,
                    "sheet": primary.sheet,
                    "row": start_row,
                    "project": primary.context.project,
                    "internal_model": primary.context.internal_model,
                    "kind": "重复BOM区段未重复汇总",
                    "detail": (
                        f"原始主料行 {start_row}-{end_row} 与较早区段 {first_start}-{first_end} 逐项相同；"
                        "标准表只保留一次，避免后续数据库重复。"
                    ),
                }
            )
            continue
        if prior and prior[0] != signature:
            primary = block[0].primary
            anomalies.append(
                {
                    "filename": primary.filename,
                    "sheet": primary.sheet,
                    "row": start_row,
                    "project": primary.context.project,
                    "internal_model": primary.context.internal_model,
                    "kind": "同型号重复区段内容不同",
                    "detail": "相同项目/内部型号再次出现但内容不同；两个区段均保留，需工程人员确认版本。",
                }
            )
        else:
            seen[key] = (signature, start_row, end_row)
        output.extend(block)
    return output, anomalies, {
        "duplicate_sections_suppressed": duplicate_sections,
        "duplicate_groups_suppressed": duplicate_groups,
    }


def template_field(value: Any) -> str | None:
    aliases = {
        "序号": "sequence",
        "项目号": "project",
        "板子类型": "board_type",
        "内部型号": "internal_model",
        "物料规格描述": "specification",
        "品牌": "brand",
        "用量": "quantity",
        "数量": "quantity",
        "替代料": "substitute",
        "供应商": "supplier",
        "订单数量": "order_quantity",
        "需求数量": "demand_quantity",
        "购买数量": "purchase_quantity",
        "库存数": "inventory",
    }
    return aliases.get(normalized_header(value))


def parse_template_target(path: Path) -> list[TemplateRow]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        if workbook.sheetnames[:2] != ["原BOM", "Sheet1"]:
            raise ValueError("TEMPLATE_SHEET_ORDER_MISMATCH")
        sheet = workbook["Sheet1"]
        header = tuple(text(sheet.cell(1, column).value) for column in range(1, len(TEMPLATE_HEADERS) + 1))
        if header != TEMPLATE_HEADERS:
            raise ValueError("TEMPLATE_HEADER_MISMATCH")
        active: dict[str, int] = {}
        rows: list[TemplateRow] = []
        for row_no, cells in enumerate(sheet.iter_rows(min_row=1, max_col=len(TEMPLATE_HEADERS)), 1):
            values = [cell.value for cell in cells]
            mapping = {field_name: index for index, value in enumerate(values) if (field_name := template_field(value))}
            if len(mapping) >= 6 and all(field_name in mapping for field_name in ("sequence", "project", "board_type", "internal_model", "specification")):
                active = mapping
                continue
            if not active:
                continue

            def get(field_name: str) -> Any:
                index = active.get(field_name)
                return values[index] if index is not None and index < len(values) else None

            project = text(get("project"))
            specification = text(get("specification"))
            if not project and not specification:
                continue
            if not project or not specification:
                raise ValueError(f"TEMPLATE_PARTIAL_DATA_ROW:{row_no}")
            rows.append(
                TemplateRow(
                    source_row=row_no,
                    project=project,
                    board_type=text(get("board_type")),
                    internal_model=text(get("internal_model")),
                    specification=specification,
                    brand=text(get("brand")),
                    quantity=excel_number(get("quantity"), positive=True),
                    substitute=text(get("substitute")),
                    supplier=text(get("supplier")),
                    order_quantity=excel_number(get("order_quantity")),
                    inventory=excel_number(get("inventory")),
                )
            )
        if len(rows) != 53:
            raise ValueError(f"TEMPLATE_DATA_ROW_COUNT_MISMATCH:{len(rows)}")
        return rows
    finally:
        workbook.close()


def template_rows_to_standard(rows: Iterable[TemplateRow], template_name: str) -> list[StandardRow]:
    return [
        StandardRow(
            project=row.project,
            board_type=row.board_type,
            internal_model=row.internal_model,
            specification=row.specification,
            brand=row.brand,
            quantity=row.quantity,
            substitute=row.substitute,
            supplier=row.supplier,
            order_quantity=row.order_quantity,
            inventory=row.inventory,
            source_file=template_name,
            source_sheet="Sheet1",
            source_row=row.source_row,
            rule="模板第二张整理结果；已由第一张原BOM行组与用量复核",
            status="模板确认标准",
        )
        for row in rows
    ]


def template_line_has_evidence(target: TemplateRow, group: RawGroup) -> bool:
    target_key = strict_key(target.specification)
    for line in (group.primary, *group.alternatives):
        for value in (line.manufacturer_part_no, line.specification, line.name):
            key = strict_key(value)
            if key and (key in target_key or target_key in key):
                return True
        standardized, _rule = standardized_specification(line)
        generic_key = strict_key(standardized)
        if generic_key and (generic_key in target_key or target_key in generic_key):
            return True
    return False


def validate_template_pair(template_path: Path) -> dict[str, Any]:
    target_rows = parse_template_target(template_path)
    raw_groups, anomalies, stats = parse_classic_groups(
        template_path,
        project="A200",
        sheet_name="原BOM",
    )
    if anomalies:
        unexpected = [item for item in anomalies if item["kind"] != "板件本体未计入"]
        if unexpected:
            raise ValueError(f"TEMPLATE_RAW_UNEXPECTED_ANOMALY:{unexpected[0]['kind']}")
    if len(raw_groups) != len(target_rows):
        raise ValueError(f"TEMPLATE_RAW_TARGET_COUNT_MISMATCH:{len(raw_groups)}:{len(target_rows)}")
    evidence_failures: list[int] = []
    quantity_failures: list[int] = []
    context_failures: list[int] = []
    for target, group in zip(target_rows, raw_groups):
        if target.quantity != group.quantity:
            quantity_failures.append(target.source_row)
        if target.project != group.primary.context.project or target.internal_model != group.primary.context.internal_model:
            context_failures.append(target.source_row)
        if not template_line_has_evidence(target, group):
            evidence_failures.append(target.source_row)
    if quantity_failures or context_failures or evidence_failures:
        raise ValueError(
            "TEMPLATE_PAIR_VALIDATION_FAILED:"
            f"quantity={quantity_failures}:context={context_failures}:evidence={evidence_failures}"
        )
    return {
        "raw_groups": len(raw_groups),
        "target_rows": len(target_rows),
        "raw_alternative_rows": stats["alternative_rows"],
        "raw_board_rows_excluded": stats["board_base_excluded"],
        "row_evidence_matches": len(target_rows),
        "quantity_matches": len(target_rows),
    }


def parse_a200_material_list(path: Path) -> ParseResult:
    workbook = sparse_xlsx_cells(path, max_column=14)
    sheet_name = "Sheet2"
    if sheet_name not in workbook:
        raise ValueError("A200_MATERIAL_LIST_SHEET_MISSING")
    contexts = {
        "8SD05169C": "USB小板",
        "3PF05170C": "主FPC",
        "4P05171B": "USB FPC",
        "2P05173D": "屏排线",
        "2P05174B": "闪光灯",
        "2P05175A": "自定义键",
        "2P05176A": "侧键",
        "2SH05177A": "光感小板",
    }
    rows: list[StandardRow] = []
    anomalies: list[dict[str, Any]] = []
    current_model = ""
    current_board = ""
    section_count = 0
    for row_no, cells in sorted(workbook[sheet_name].items()):
        if row_no <= 3:
            continue
        model = text(cells.get(3))
        if model in contexts and text(cells.get(2)):
            current_model = model
            current_board = contexts[model]
            section_count += 1
            continue
        if row_no >= 69 or not current_model:
            continue
        name = text(cells.get(2))
        specification = text(cells.get(4))
        supplier = text(cells.get(5))
        if not name or normalized_header(name) in {"制作", "审核", "合计"}:
            continue
        if any(marker in name for marker in ("订单日期", "交货日期")):
            continue
        standard_spec = tidy_spec(f"{name},{specification}" if specification else name)
        rows.append(
            StandardRow(
                project="A200",
                board_type=current_board,
                internal_model=current_model,
                specification=standard_spec,
                brand="",
                quantity=None,
                substitute="",
                supplier=supplier,
                source_file=path.name,
                source_sheet=sheet_name,
                source_row=row_no,
                rule="物料清单名称+明确规格；供应商仅取原表供应商列；用量缺失留空",
                status="用量待补",
            )
        )
        anomalies.append(
            {
                "filename": path.name,
                "sheet": sheet_name,
                "row": row_no,
                "project": "A200",
                "internal_model": current_model,
                "kind": "用量待确认",
                "detail": "量产物料清单未提供单机用量；输出用量留空。",
            }
        )
    return ParseResult(
        rows=rows,
        anomalies=anomalies,
        stats={"source_nonempty_rows": len(workbook[sheet_name]), "sections": section_count, "groups": len(rows)},
    )


def parse_source(
    path: Path,
    *,
    project: str,
    sheet_name: str,
    fallback_model: str = "",
    special_j587: bool = False,
) -> ParseResult:
    groups, anomalies, stats = parse_classic_groups(
        path,
        project=project,
        sheet_name=sheet_name,
        fallback_model=fallback_model,
        special_j587=special_j587,
    )
    groups, duplicate_anomalies, duplicate_stats = suppress_identical_repeated_sections(groups)
    anomalies.extend(duplicate_anomalies)
    stats.update(duplicate_stats)
    stats["groups_after_exact_section_dedup"] = len(groups)
    rows = standard_rows_from_groups(groups)
    for row in rows:
        if not row.board_type:
            anomalies.append(
                {
                    "filename": row.source_file,
                    "sheet": row.source_sheet,
                    "row": row.source_row,
                    "project": row.project,
                    "internal_model": row.internal_model,
                    "kind": "板子类型待确认",
                    "detail": "原表分段只给出内部型号，未给出可证明的板子类型；标准表保持空白。",
                }
            )
        if not row.internal_model:
            anomalies.append(
                {
                    "filename": row.source_file,
                    "sheet": row.source_sheet,
                    "row": row.source_row,
                    "project": row.project,
                    "internal_model": "",
                    "kind": "内部型号待确认",
                    "detail": "原表标题未给出可证明的内部型号；标准表保持空白。",
                }
            )
    return ParseResult(rows=rows, anomalies=anomalies, stats=stats)


def compare_a200_legacy(template_rows: list[TemplateRow], legacy_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    legacy_groups, legacy_anomalies, legacy_stats = parse_classic_groups(
        legacy_path,
        project="A200",
        sheet_name="BOM",
    )
    if len(legacy_groups) != len(template_rows):
        raise ValueError(f"A200_LEGACY_GROUP_COUNT_MISMATCH:{len(legacy_groups)}:{len(template_rows)}")
    conflicts: list[dict[str, Any]] = []
    for target, group in zip(template_rows, legacy_groups):
        evidence = template_line_has_evidence(target, group)
        quantity_equal = target.quantity == group.quantity
        if evidence and quantity_equal:
            continue
        conflicts.append(
            {
                "filename": legacy_path.name,
                "sheet": "BOM",
                "row": group.primary.row_no,
                "project": "A200",
                "internal_model": target.internal_model,
                "kind": "模板与旧A200版本差异",
                "detail": (
                    f"模板标准行 {target.source_row} 优先；"
                    f"规格证据一致={str(evidence).lower()}，用量一致={str(quantity_equal).lower()}。"
                ),
            }
        )
    legacy_board_exclusions = [item for item in legacy_anomalies if item["kind"] == "板件本体未计入"]
    unexpected = [item for item in legacy_anomalies if item["kind"] != "板件本体未计入"]
    if unexpected:
        raise ValueError(f"A200_LEGACY_UNEXPECTED_ANOMALY:{unexpected[0]['kind']}")
    return conflicts, {
        "legacy_groups": len(legacy_groups),
        "legacy_alternative_rows": legacy_stats["alternative_rows"],
        "legacy_board_rows_excluded": len(legacy_board_exclusions),
        "template_precedence_conflicts": len(conflicts),
    }


def build_data(source_dir: Path, template_path: Path) -> dict[str, Any]:
    template_rows = parse_template_target(template_path)
    template_validation = validate_template_pair(template_path)
    sheet_rows: dict[str, list[StandardRow]] = {}
    anomalies: list[dict[str, Any]] = []
    source_stats: dict[str, dict[str, Any]] = {}

    source_specs = (
        ("1928C标准", "1928C量产BOM.xlsx", "1928C", "M105H-DG-M23U-charge-V1_BOM", "SC6501_SUB_V1.0.1", False),
        ("A118标准", "A118量产BOM.xlsx", "A118", "SHEET1", "", False),
        ("G20-G15G标准", "G20-G15G项目量产BOM.xlsx", "G20-G15G", "G9_5G_ANT_SCH_V2R0", "", False),
        ("J587标准", "J587_SUBA2_V01-20260703.xlsx", "J587", "TYPE-C耳机小板", "", True),
        ("V700标准", "V700量产BOM.xlsx", "V700", "BOM", "", False),
    )
    for output_sheet, filename, project, source_sheet, fallback_model, special_j587 in source_specs:
        result = parse_source(
            source_dir / filename,
            project=project,
            sheet_name=source_sheet,
            fallback_model=fallback_model,
            special_j587=special_j587,
        )
        sheet_rows[output_sheet] = result.rows
        anomalies.extend(result.anomalies)
        source_stats[filename] = result.stats | {"standard_rows": len(result.rows), "handling": "PARSED_TO_STANDARD"}

    a200_rows = template_rows_to_standard(template_rows, template_path.name)
    sheet_rows["A200量产BOM标准"] = a200_rows
    a200_conflicts, a200_stats = compare_a200_legacy(template_rows, source_dir / "A200量产BOM.xlsx")
    anomalies.extend(a200_conflicts)
    source_stats["A200量产BOM.xlsx"] = a200_stats | {
        "standard_rows": len(a200_rows),
        "handling": "SUPERSEDED_BY_MOBAN_CANONICAL_PAIR",
    }

    material_list = parse_a200_material_list(source_dir / "A200量产物料清单.xlsx")
    sheet_rows["A200物料清单标准"] = material_list.rows
    anomalies.extend(material_list.anomalies)
    source_stats["A200量产物料清单.xlsx"] = material_list.stats | {
        "standard_rows": len(material_list.rows),
        "handling": "PARSED_TO_STANDARD_WITH_UNKNOWN_QUANTITY",
    }

    sheet_rows["A200量产注意事项"] = []
    source_stats["A200量产注意事项.xls"] = {
        "standard_rows": 0,
        "handling": "ARCHIVE_NOTE_NO_MATERIAL_ROWS",
    }
    anomalies.append(
        {
            "filename": "A200量产注意事项.xls",
            "sheet": "Sheet1",
            "row": 0,
            "project": "A200",
            "internal_model": "",
            "kind": "说明档不计入汇总",
            "detail": "该 XLS 是量产注意事项/说明档，不含可转换物料行；保留空标准页和来源记录。",
        }
    )

    title_model = model_from_title("J587_SUBA1_V01_20251215", project="J587")
    filename_model_match = J587_MODEL_RE.search(Path("J587_SUBA2_V01-20260703.xlsx").stem.upper())
    filename_model = filename_model_match.group(1).replace("-", "_") if filename_model_match else ""
    if filename_model and title_model and filename_model != title_model:
        anomalies.append(
            {
                "filename": "J587_SUBA2_V01-20260703.xlsx",
                "sheet": "TYPE-C耳机小板",
                "row": 1,
                "project": "J587",
                "internal_model": title_model,
                "kind": "文件名与表内版本冲突",
                "detail": f"文件名标识 {filename_model}，表内标题标识 {title_model}；按原表标题填写并保留人工确认。",
            }
        )

    missing_source_stats = EXPECTED_SOURCE_FILES - set(source_stats)
    if missing_source_stats:
        raise ValueError(f"SOURCE_NOT_ACCOUNTED_FOR:{sorted(missing_source_stats)}")
    ordered_rows = {sheet: sheet_rows.get(sheet, []) for sheet in SOURCE_SHEET_ORDER}
    consolidated = [row for sheet in SOURCE_SHEET_ORDER for row in ordered_rows[sheet]]
    if not consolidated:
        raise ValueError("STANDARD_ROWS_EMPTY")
    return {
        "sheets": ordered_rows,
        "consolidated": consolidated,
        "anomalies": anomalies,
        "source_stats": source_stats,
        "template_validation": template_validation,
        "counts": {
            "standard_rows": len(consolidated),
            "source_standard_sheets": len(SOURCE_SHEET_ORDER),
            "nonempty_source_standard_sheets": sum(bool(rows) for rows in ordered_rows.values()),
            "anomalies": len(anomalies),
            "unknown_quantity_rows": sum(row.quantity is None for row in consolidated),
            "alternative_bearing_rows": sum(bool(row.substitute) for row in consolidated),
        },
    }


def _copy_template_header_style(template_path: Path) -> tuple[list[Any], dict[str, float], float | None]:
    workbook = openpyxl.load_workbook(template_path, read_only=False, data_only=False, keep_links=False)
    try:
        sheet = workbook["Sheet1"]
        styles = []
        for column in range(1, len(TEMPLATE_HEADERS) + 1):
            cell = sheet.cell(1, column)
            styles.append(
                {
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "protection": copy(cell.protection),
                }
            )
        widths = {
            get_column_letter(column): sheet.column_dimensions[get_column_letter(column)].width
            for column in range(1, len(TEMPLATE_HEADERS) + 1)
        }
        return styles, widths, sheet.row_dimensions[1].height
    finally:
        workbook.close()


def apply_header_style(sheet: Any, header_styles: list[Any], header_height: float | None) -> None:
    for column, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = sheet.cell(1, column, header)
        style = header_styles[column - 1]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = style["number_format"]
        cell.protection = copy(style["protection"])
    sheet.row_dimensions[1].height = header_height or 22


def standard_cell_values(row: StandardRow, excel_row: int, sequence: int) -> list[Any]:
    return [
        sequence,
        safe_excel_text(row.project),
        safe_excel_text(row.board_type),
        safe_excel_text(row.internal_model),
        safe_excel_text(row.specification),
        safe_excel_text(row.brand),
        row.quantity,
        safe_excel_text(row.substitute),
        safe_excel_text(row.supplier),
        row.order_quantity,
        f'=IF(OR(G{excel_row}="",J{excel_row}=""),"",G{excel_row}*J{excel_row})',
        f'=IF(OR(K{excel_row}="",M{excel_row}=""),"",MAX(K{excel_row}-M{excel_row},0))',
        row.inventory,
    ]


def write_standard_sheet(
    workbook: Workbook,
    title: str,
    rows: list[StandardRow],
    header_styles: list[Any],
    widths: dict[str, float],
    header_height: float | None,
    *,
    consolidated: bool = False,
) -> None:
    sheet = workbook.create_sheet(title)
    apply_header_style(sheet, header_styles, header_height)
    thin = Side(style="thin", color="FFD9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    group_fill = PatternFill("solid", fgColor="FFF3F6FA")
    previous_context: tuple[str, str, str] | None = None
    for sequence, row in enumerate(rows, 1):
        excel_row = sequence + 1
        values = standard_cell_values(row, excel_row, sequence)
        context = (row.project, row.board_type, row.internal_model)
        for column, value in enumerate(values, 1):
            cell = sheet.cell(excel_row, column, value)
            cell.font = Font(name="宋体", size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if column in (1, 2, 3, 4, 7, 10, 11, 12, 13) else "left",
                vertical="center",
                wrap_text=True,
            )
            if context != previous_context:
                cell.fill = group_fill
        source_label = f"来源：{row.source_file} / {row.source_sheet} / 主料行 {row.source_row}"
        if row.alternative_rows:
            source_label += " / 备选行 " + ",".join(str(value) for value in row.alternative_rows)
        sheet.cell(excel_row, 1).comment = Comment(source_label, "Codex")
        sheet.row_dimensions[excel_row].height = 30
        previous_context = context
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.column_dimensions["E"].width = max(widths.get("E", 50), 52)
    sheet.column_dimensions["H"].width = max(widths.get("H", 18), 28)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{max(1, len(rows) + 1)}"
    sheet.sheet_view.showGridLines = False
    if not rows:
        sheet.sheet_properties.tabColor = "FF808080"
        sheet.cell(2, 5, "本来源为说明档，无可转换物料行；未计入全部物料汇总。")
        sheet.cell(2, 5).font = Font(name="宋体", size=10, italic=True, color="FF666666")
        sheet.cell(2, 5).alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[2].height = 32
    elif consolidated:
        sheet.sheet_properties.tabColor = "FF4472C4"


def write_provenance_sheet(workbook: Workbook, data: dict[str, Any]) -> None:
    headers = (
        "汇总序号",
        "标准Sheet",
        "标准行号",
        "源文件",
        "源Sheet",
        "原始主料行",
        "原始备选料行",
        "项目号",
        "板子类型",
        "内部型号",
        "整理规则",
        "状态",
    )
    sheet = workbook.create_sheet("来源追溯")
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(name="宋体", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    global_sequence = 0
    row_no = 2
    for standard_sheet in SOURCE_SHEET_ORDER:
        for local_sequence, row in enumerate(data["sheets"][standard_sheet], 1):
            global_sequence += 1
            values = (
                global_sequence,
                standard_sheet,
                local_sequence + 1,
                row.source_file,
                row.source_sheet,
                row.source_row,
                ",".join(str(value) for value in row.alternative_rows),
                row.project,
                row.board_type,
                row.internal_model,
                row.rule,
                row.status,
            )
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_no, column, safe_excel_text(value) if isinstance(value, str) else value)
                cell.font = Font(name="宋体", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_no += 1
    widths = (10, 22, 10, 34, 24, 12, 22, 12, 18, 20, 48, 18)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{max(1, row_no - 1)}"


def write_anomaly_sheet(workbook: Workbook, anomalies: list[dict[str, Any]]) -> None:
    headers = ("源文件", "源Sheet", "原始行", "项目号", "内部型号", "类型", "说明")
    sheet = workbook.create_sheet("整理异常")
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(name="宋体", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFC65911")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_no, item in enumerate(anomalies, 2):
        values = (
            item["filename"],
            item["sheet"],
            item["row"],
            item["project"],
            item["internal_model"],
            item["kind"],
            item["detail"],
        )
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_no, column, safe_excel_text(value) if isinstance(value, str) else value)
            cell.font = Font(name="宋体", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = (34, 24, 10, 12, 20, 24, 65)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(1, len(anomalies) + 1)}"


def write_source_notes_sheet(
    workbook: Workbook,
    data: dict[str, Any],
    source_manifest: dict[str, Any],
    template_snapshot: dict[str, Any],
) -> None:
    headers = ("文件", "SHA-256", "处理方式", "标准行数", "说明")
    sheet = workbook.create_sheet("来源与说明")
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(name="宋体", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF548235")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    manifest_by_file = {item["filename"]: item for item in source_manifest["files"]}
    handling_text = {
        "PARSED_TO_STANDARD": "按原表主料/备选料组整理；主料一行，备选料折叠到替代料。",
        "SUPERSEDED_BY_MOBAN_CANONICAL_PAIR": "与模板为同一 A200 逻辑 BOM；存在版本差异，按项目负责人指定的 moban 原BOM/Sheet1 为准，不重复汇总。",
        "PARSED_TO_STANDARD_WITH_UNKNOWN_QUANTITY": "按板型/内部型号整理名称、规格和明确供应商；原表无单机用量，留空。",
        "ARCHIVE_NOTE_NO_MATERIAL_ROWS": "量产说明档，无可转换物料行；保留空标准页，不计入汇总。",
    }
    row_no = 2
    for filename in sorted(EXPECTED_SOURCE_FILES):
        stats = data["source_stats"][filename]
        values = (
            filename,
            manifest_by_file[filename]["sha256"],
            stats["handling"],
            stats.get("standard_rows", 0),
            handling_text[stats["handling"]],
        )
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_no, column, safe_excel_text(value) if isinstance(value, str) else value)
            cell.font = Font(name="宋体", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_no += 1
    template_values = (
        template_snapshot["filename"],
        template_snapshot["sha256"],
        "RAW_TO_STANDARD_TEMPLATE",
        data["template_validation"]["target_rows"],
        "第一张原BOM是转换依据，第二张Sheet1是13列目标；53个主料组逐行规格证据和用量均已核对。",
    )
    for column, value in enumerate(template_values, 1):
        cell = sheet.cell(row_no, column, safe_excel_text(value) if isinstance(value, str) else value)
        cell.font = Font(name="宋体", size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    row_no += 2
    notes = (
        "汇总原则：保留各板型BOM用量行，不跨文件模糊去重；相同物料在不同板型出现时保留多行。",
        "工作表范围：每个文件以当前BOM/物料主表为准；更新记录、变更记录及空Sheet3属于历史或空页，不作为当前采购明细重复汇总。",
        "替代料原则：除模板已人工选定项外，其他来源以原表主料为本行，后续备选型号折叠到替代料列。",
        "供应商原则：仅模板第二张和A200量产物料清单的明确供应商列可填写；联系人/备注不当作供应商。",
        "数量原则：订单数量、库存数未知时留空；需求/购买数量为可审阅公式，不伪造零值。",
        "数据库边界：本工作簿尚未写入数据库；正式导入、去重、审核和编码必须另行授权。",
    )
    for note in notes:
        sheet.cell(row_no, 1, "规则")
        sheet.cell(row_no, 2, note)
        sheet.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=5)
        sheet.cell(row_no, 1).font = Font(name="宋体", size=9, bold=True)
        sheet.cell(row_no, 2).font = Font(name="宋体", size=9)
        sheet.cell(row_no, 2).alignment = Alignment(wrap_text=True, vertical="top")
        row_no += 1
    widths = (34, 68, 38, 12, 72)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"


def write_workbook(
    output: Path,
    data: dict[str, Any],
    template_path: Path,
    source_manifest: dict[str, Any],
    template_snapshot: dict[str, Any],
) -> None:
    header_styles, widths, header_height = _copy_template_header_style(template_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    write_standard_sheet(
        workbook,
        "全部物料汇总",
        data["consolidated"],
        header_styles,
        widths,
        header_height,
        consolidated=True,
    )
    for title in SOURCE_SHEET_ORDER:
        write_standard_sheet(
            workbook,
            title,
            data["sheets"][title],
            header_styles,
            widths,
            header_height,
        )
    write_provenance_sheet(workbook, data)
    write_anomaly_sheet(workbook, data["anomalies"])
    write_source_notes_sheet(workbook, data, source_manifest, template_snapshot)
    workbook.properties.creator = "晨亿达 ERP / Codex"
    workbook.properties.title = "全项目标准物料明细"
    workbook.properties.description = "按 moban.xlsx 原BOM到Sheet1规则逐表标准化；未写数据库"
    output.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    with tempfile.NamedTemporaryFile(prefix=f".{output.name}.", suffix=".tmp", dir=output.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.chmod(temporary, 0o600)
        os.replace(temporary, output)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def validate_workbook(output: Path, data: dict[str, Any]) -> dict[str, Any]:
    with ZipFile(output) as archive:
        bad_zip_member = archive.testzip()
        names = set(archive.namelist())
        if bad_zip_member:
            raise ValueError(f"OUTPUT_ZIP_INVALID:{bad_zip_member}")
        if any(name.endswith("vbaProject.bin") or name.startswith("xl/externalLinks/") for name in names):
            raise ValueError("OUTPUT_UNSAFE_FEATURE")
    workbook = openpyxl.load_workbook(output, read_only=False, data_only=False, keep_links=False)
    try:
        expected_sheets = [
            "全部物料汇总",
            *SOURCE_SHEET_ORDER,
            "来源追溯",
            "整理异常",
            "来源与说明",
        ]
        if workbook.sheetnames != expected_sheets:
            raise ValueError("OUTPUT_SHEET_ORDER_MISMATCH")
        standard_sheets = ["全部物料汇总", *SOURCE_SHEET_ORDER]
        total_formula_count = 0
        phone_hits = 0
        sensitive_hits = 0
        for title in standard_sheets:
            sheet = workbook[title]
            headers = tuple(text(sheet.cell(1, column).value) for column in range(1, len(TEMPLATE_HEADERS) + 1))
            if headers != TEMPLATE_HEADERS:
                raise ValueError(f"OUTPUT_HEADER_MISMATCH:{title}")
            expected_rows = len(data["consolidated"]) if title == "全部物料汇总" else len(data["sheets"][title])
            for row_no in range(2, expected_rows + 2):
                quantity = sheet.cell(row_no, 7).value
                if quantity is not None and excel_number(quantity, positive=True) is None:
                    raise ValueError(f"OUTPUT_QUANTITY_INVALID:{title}:{row_no}")
                demand = text(sheet.cell(row_no, 11).value)
                purchase = text(sheet.cell(row_no, 12).value)
                expected_demand = f'=IF(OR(G{row_no}="",J{row_no}=""),"",G{row_no}*J{row_no})'
                expected_purchase = f'=IF(OR(K{row_no}="",M{row_no}=""),"",MAX(K{row_no}-M{row_no},0))'
                if demand != expected_demand or purchase != expected_purchase:
                    raise ValueError(f"OUTPUT_FORMULA_INVALID:{title}:{row_no}")
                total_formula_count += 2
            expected_physical_rows = expected_rows + 1 if expected_rows else 2
            if sheet.max_row != expected_physical_rows:
                raise ValueError(f"OUTPUT_ROW_COUNT_MISMATCH:{title}:{sheet.max_row}:{expected_physical_rows}")
        if workbook["全部物料汇总"].max_row != data["counts"]["standard_rows"] + 1:
            raise ValueError("OUTPUT_CONSOLIDATED_COUNT_MISMATCH")
        if workbook["来源追溯"].max_row != data["counts"]["standard_rows"] + 1:
            raise ValueError("OUTPUT_PROVENANCE_COUNT_MISMATCH")
        if workbook["整理异常"].max_row != data["counts"]["anomalies"] + 1:
            raise ValueError("OUTPUT_ANOMALY_COUNT_MISMATCH")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = text(cell.value)
                    if not value:
                        continue
                    phone_hits += len(PHONE_LIKE_RE.findall(value))
                    sensitive_hits += len(SENSITIVE_RE.findall(value))
                    if value.startswith("=") and cell.column not in (11, 12):
                        raise ValueError(f"OUTPUT_UNEXPECTED_FORMULA:{sheet.title}:{cell.coordinate}")
        if phone_hits or sensitive_hits:
            raise ValueError(f"OUTPUT_SENSITIVE_CONTENT:phone={phone_hits}:secret={sensitive_hits}")
        return {
            "zip_integrity": "PASS",
            "macro_external_links": 0,
            "sheet_contract": "PASS",
            "standard_rows": data["counts"]["standard_rows"],
            "provenance_rows": data["counts"]["standard_rows"],
            "formula_cells": total_formula_count,
            "phone_like_values": phone_hits,
            "sensitive_values": sensitive_hits,
        }
    finally:
        workbook.close()


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", prefix=f".{path.name}.", suffix=".tmp", dir=path.parent, delete=False
    ) as handle:
        temporary = Path(handle.name)
        json.dump(payload, handle, ensure_ascii=False, sort_keys=True, indent=2)
        handle.write("\n")
    try:
        os.chmod(temporary, 0o600)
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def export_standardized_workbook(
    *,
    source_dir: Path,
    template_path: Path,
    manifest_json: Path,
    output: Path,
    report_path: Path,
    confirmation: str,
) -> dict[str, Any]:
    if confirmation != CONFIRMATION:
        raise ValueError("CONFIRMATION_REQUIRED")
    if output.suffix.lower() != ".xlsx" or output.resolve() == template_path.resolve():
        raise ValueError("OUTPUT_PATH_INVALID")
    os.umask(0o077)
    source_manifest = json.loads(manifest_json.read_text(encoding="utf-8"))
    source_before = verify_manifest(source_dir, source_manifest)
    template_before = file_snapshot(template_path)
    data = build_data(source_dir, template_path)
    write_workbook(output, data, template_path, source_manifest, template_before)
    validation = validate_workbook(output, data)
    source_after = verify_manifest(source_dir, source_manifest)
    template_after = file_snapshot(template_path)
    if source_before != source_after or template_before != template_after:
        raise ValueError("INPUT_CHANGED_DURING_EXPORT")
    mode = stat.S_IMODE(output.stat().st_mode)
    if mode != 0o600:
        raise ValueError("OUTPUT_MODE_INVALID")
    sheet_counts = {sheet: len(rows) for sheet, rows in data["sheets"].items()}
    report = {
        "task": TASK,
        "result": "OFFLINE_STANDARDIZED_BOM_WORKBOOK_CREATED_REVIEW_REQUIRED",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_manifest_sha256": source_manifest["manifest_sha256"],
        "template": {
            "filename": template_before["filename"],
            "sha256": template_before["sha256"],
            "raw_sheet": "原BOM",
            "standard_sheet": "Sheet1",
            "headers": list(TEMPLATE_HEADERS),
            "validation": data["template_validation"],
        },
        "output": {
            "path": str(output),
            "sha256": sha256_path(output),
            "size_bytes": output.stat().st_size,
            "mode": oct(mode),
        },
        "counts": data["counts"] | {"by_sheet": sheet_counts},
        "source_stats": data["source_stats"],
        "anomaly_types": dict(Counter(item["kind"] for item in data["anomalies"])),
        "validation": validation | {"source_files_unchanged": "PASS", "template_unchanged": "PASS"},
        "database_writes": 0,
        "deployment_changes": 0,
    }
    atomic_write_json(report_path, report)
    return report


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--manifest-json", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--confirm", required=True)
    args = parser.parse_args()
    report = export_standardized_workbook(
        source_dir=args.source_dir,
        template_path=args.template,
        manifest_json=args.manifest_json,
        output=args.output,
        report_path=args.report,
        confirmation=args.confirm,
    )
    print(
        json.dumps(
            {
                "result": report["result"],
                "output": report["output"],
                "counts": report["counts"],
                "validation": report["validation"],
                "anomaly_types": report["anomaly_types"],
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
