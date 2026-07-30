#!/usr/bin/env python3
"""Build a reviewable internal-material workbook from guarded BOM evidence.

This exporter is intentionally offline.  It reads the immutable source files and
the root-only evidence produced by SELFHOST-LANDING-TASK02, creates an XLSX
workbook, and never connects to a database or assigns a new formal ERP code.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import posixpath
import re
import stat
import tempfile
import unicodedata
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TASK = "SELFHOST-LANDING-TASK06"
CONFIRMATION = "OFFLINE_INTERNAL_MATERIAL_LIBRARY_EXPORT"
TEMPLATE_HEADERS = (
    "序号",
    "项目号",
    "板子类型",
    "内部型号",
    "物料规格描述",
    "品牌",
    "用量",
    "替代料",
    "供应商",
    "订单数量",
    "需求数量",
    "购买数量",
    "库存数",
)
LIBRARY_HEADERS = (
    "内部记录ID",
    "正式ERP编码",
    "记录状态",
    "物料分类",
    "标准物料名称",
    "标准规格描述",
    "品牌/生产厂商",
    "制造商料号/型号",
    "封装",
    "单位",
    "来源项目",
    "来源文件数",
    "来源行数",
    "身份依据",
    "待确认事项",
)
REVIEW_HEADERS = (
    "待确认ID",
    "来源引用",
    "项目号",
    "问题层级",
    "原因",
    "建议处理",
    "正式ERP编码",
    "物料规格描述",
    "品牌",
    "源文件",
    "Sheet",
    "原始行号",
)
SOURCE_HEADERS = (
    "内部记录ID",
    "正式ERP编码",
    "来源引用",
    "来源类型",
    "项目号",
    "源文件",
    "Sheet",
    "原始行号",
    "来源SHA前12位",
    "映射状态",
    "映射依据",
    "标准明细序号",
)

REASON_TEXT = {
    "BOM_QUANTITY_NOT_VALID": "BOM 用量缺失、非正数或不能从位号安全推导",
    "CATEGORY_OR_COUNT_UNIT_NOT_DETERMINISTIC": "物料类别或可数件单位不能确定",
    "STABLE_IDENTITY_CONFLICT": "同一稳定身份对应的规格证据互相冲突",
    "STABLE_MATERIAL_IDENTITY_MISSING": "缺少可验证的稳定物料身份",
    "ARCHIVAL_PROCESS_OR_CHANGE_NOTE": "流程/变更说明，仅归档",
    "NON_DATA_NOTE_OR_FOOTER": "非物料数据说明或页脚",
    "REPEATED_HEADER": "重复表头",
    "TEMPLATE_STRICT_MATCH_MISSING": "模板行无法与单一既有物料严格对应",
    "TEMPLATE_STRICT_MATCH_AMBIGUOUS": "模板行严格命中多个既有物料编码",
    "SOURCE_CONTEXT_VERSION_CONFLICT": "文件名与表内标题的板卡版本标识冲突",
}
IDENTITY_TEXT = {
    "SOURCE_CODE": "来源稳定编码",
    "MPN": "制造商 + 完整 MPN",
    "STRICT_SPEC_COMPOSITE": "类别 + 完整规格组合",
    "": "来源行隔离；尚无安全归并依据",
}
PROJECT_PREFIXES = (
    ("1928C", "1928C"),
    ("A118", "A118"),
    ("A200", "A200"),
    ("G20-G15G", "G20-G15G"),
    ("J587", "J587"),
    ("V700", "V700"),
)

PHONE_LIKE_RE = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")
SENSITIVE_RE = re.compile(
    r"(?:postgres(?:ql)?://|database_url|password\s*[=:]|api[_-]?key\s*[=:]|bearer\s+[a-z0-9._-]+)",
    re.I,
)
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
REF_RE = re.compile(r"(?<![A-Z0-9])([A-Z]{1,3}\d+)(?![A-Z0-9])")

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


@dataclass(frozen=True)
class TemplateRow:
    source_row: int
    project: str
    board_type: str
    internal_model: str
    specification: str
    brand: str
    quantity: int | float | None
    substitute: str
    supplier: str
    order_quantity: int | float | None
    inventory: int | float | None


@dataclass(frozen=True)
class Context:
    project: str
    board_type: str
    internal_model: str
    title: str


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return unicodedata.normalize("NFC", str(value)).strip()


def normalized_header(value: Any) -> str:
    return re.sub(r"[\s\-_/\\:：()（）\[\]【】.]+", "", text(value).lower())


def strict_key(value: Any) -> str:
    value_text = unicodedata.normalize("NFKC", text(value)).upper()
    value_text = value_text.translate(str.maketrans({"，": ",", "；": ";", "：": ":", "。": "."}))
    value_text = re.sub(r"\s+", "", value_text)
    return value_text.strip(",;")


def safe_excel_text(value: Any, *, limit: int = 30000, redact_phone: bool = True) -> str:
    result = CONTROL_RE.sub("", text(value))
    if redact_phone:
        result = PHONE_LIKE_RE.sub("[已移除可能的联系电话]", result)
    if result.lstrip().startswith(("=", "+", "-", "@")):
        result = "'" + result
    return result[:limit]


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_sha(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def stable_candidate_id(prefix: str, value: str) -> str:
    digest = hashlib.sha256(value.encode("utf-8")).hexdigest()[:16].upper()
    return f"{prefix}-{digest}"


def project_for(filename: str) -> str:
    stem = Path(filename).stem.upper()
    for prefix, project in PROJECT_PREFIXES:
        if stem.startswith(prefix.upper()):
            return project
    return Path(filename).stem[:80]


def excel_number(value: Any, *, positive: bool = False) -> int | float | None:
    raw = text(value).replace(",", "")
    if not raw or raw.startswith("="):
        return None
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return None
    if not number.is_finite() or number < 0 or (positive and number <= 0):
        return None
    if -number.as_tuple().exponent > 6:
        return None
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def explicit_reference_quantity(value: Any) -> int | None:
    source = unicodedata.normalize("NFKC", text(value)).upper()
    if not source or any(marker in source for marker in ("-", "~", "至")):
        return None
    tokens = REF_RE.findall(source)
    remainder = REF_RE.sub("", source)
    remainder = re.sub(r"[\s,，;；、/]+", "", remainder)
    if not tokens or remainder or len(tokens) != len(set(tokens)):
        return None
    return len(tokens)


def source_description(row: dict[str, str]) -> str:
    values = []
    seen = set()
    for field in ("specification", "model", "manufacturer_part_no", "material_name"):
        value = text(row.get(field))
        key = strict_key(value)
        if value and key not in seen:
            seen.add(key)
            values.append(value)
    return "；".join(values)[:1000] or "待补充：来源未提供物料名称/规格/型号/MPN"


def source_identity_keys(row: dict[str, str]) -> set[str]:
    return {
        key
        for key in (
            strict_key(row.get("specification")),
            strict_key(row.get("model")),
            strict_key(row.get("manufacturer_part_no")),
        )
        if key
    }


def file_snapshot(path: Path) -> dict[str, Any]:
    info = path.lstat()
    if not stat.S_ISREG(info.st_mode) or path.is_symlink():
        raise ValueError(f"INPUT_NOT_REGULAR:{path.name}")
    return {
        "filename": path.name,
        "sha256": sha256_path(path),
        "size_bytes": info.st_size,
        "inode": info.st_ino,
        "mode": oct(stat.S_IMODE(info.st_mode)),
        "uid": info.st_uid,
        "gid": info.st_gid,
        "mtime_ns": info.st_mtime_ns,
    }


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [{key: text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def verify_manifest(source_dir: Path, manifest: dict[str, Any]) -> dict[str, dict[str, Any]]:
    files = manifest.get("files")
    if not isinstance(files, list) or not files:
        raise ValueError("SOURCE_MANIFEST_EMPTY")
    if canonical_sha(files) != manifest.get("manifest_sha256"):
        raise ValueError("SOURCE_MANIFEST_DIGEST_MISMATCH")
    snapshots: dict[str, dict[str, Any]] = {}
    for expected in files:
        filename = text(expected.get("filename"))
        path = source_dir / filename
        if not path.exists():
            raise ValueError(f"SOURCE_FILE_MISSING:{filename}")
        actual = file_snapshot(path)
        for key in ("filename", "sha256", "size_bytes", "inode", "mode", "uid", "gid", "mtime_ns"):
            if str(actual[key]) != str(expected.get(key)):
                raise ValueError(f"SOURCE_FILE_DRIFT:{filename}:{key}")
        snapshots[filename] = actual
    return snapshots


def template_field(value: Any) -> str | None:
    token = normalized_header(value)
    aliases = {
        "序号": "sequence",
        "项目号": "project",
        "板子类型": "board_type",
        "内部型号": "internal_model",
        "物料规格描述": "specification",
        "品牌": "brand",
        "用量": "quantity",
        "数量": "quantity",
        "替代料": "substitute",
        "供应商": "supplier",
        "订单数量": "order_quantity",
        "需求数量": "demand_quantity",
        "购买数量": "purchase_quantity",
        "库存数": "inventory",
    }
    return aliases.get(token)


def parse_template(path: Path) -> list[TemplateRow]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        if "Sheet1" not in workbook.sheetnames:
            raise ValueError("TEMPLATE_SHEET1_MISSING")
        sheet = workbook["Sheet1"]
        header = tuple(text(sheet.cell(1, column).value) for column in range(1, len(TEMPLATE_HEADERS) + 1))
        if header != TEMPLATE_HEADERS:
            raise ValueError("TEMPLATE_HEADER_MISMATCH")
        active: dict[str, int] = {}
        rows: list[TemplateRow] = []
        for row_no, cells in enumerate(sheet.iter_rows(min_row=1, max_col=len(TEMPLATE_HEADERS)), 1):
            values = [cell.value for cell in cells]
            mapping = {field: index for index, value in enumerate(values) if (field := template_field(value))}
            if len(mapping) >= 6 and all(field in mapping for field in ("sequence", "project", "board_type", "internal_model", "specification")):
                active = mapping
                continue
            if not active:
                continue
            get = lambda field: values[active[field]] if field in active and active[field] < len(values) else None
            project = text(get("project"))
            specification = text(get("specification"))
            if not project and not specification:
                continue
            if not project or not specification:
                raise ValueError(f"TEMPLATE_PARTIAL_DATA_ROW:{row_no}")
            rows.append(
                TemplateRow(
                    source_row=row_no,
                    project=project,
                    board_type=text(get("board_type")),
                    internal_model=text(get("internal_model")),
                    specification=specification,
                    brand=text(get("brand")),
                    quantity=excel_number(get("quantity"), positive=True),
                    substitute=text(get("substitute")),
                    supplier=text(get("supplier")),
                    order_quantity=excel_number(get("order_quantity")),
                    inventory=excel_number(get("inventory")),
                )
            )
        if not rows:
            raise ValueError("TEMPLATE_DATA_ROWS_MISSING")
        return rows
    finally:
        workbook.close()


HEADER_ALIASES = {
    "sequence": ("序号", "项次", "item", "no", "no."),
    "code": ("物料编码", "内部物料编码", "hc_code", "编码"),
    "name": ("物料名称", "名称", "品名"),
    "spec": ("物料规格描述", "物料描述", "规格描述", "规格", "description"),
    "model": ("物料型号", "型号", "供应商料号", "vendorcode"),
    "maker": ("生产厂商", "制造商", "厂家", "品牌", "vendor"),
    "quantity": ("单机用量", "用量", "数量", "qty", "quantity"),
    "reference": ("位号", "位置", "reference", "ref"),
}


def generic_header_field(value: str) -> str | None:
    token = normalized_header(value)
    if not token:
        return None
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            candidate = normalized_header(alias)
            if token == candidate or (len(candidate) >= 3 and candidate in token):
                return field
    return None


def column_number(reference: str) -> int:
    match = re.match(r"([A-Z]+)", reference)
    if not match:
        return 0
    result = 0
    for char in match.group(1):
        result = result * 26 + ord(char) - 64
    return result


def sparse_xlsx_rows(path: Path) -> dict[str, dict[int, list[str]]]:
    result: dict[str, dict[int, list[str]]] = {}
    with ZipFile(path) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared = ["".join(node.text or "" for node in item.iter(f"{{{NS_MAIN}}}t")) for item in root]
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {node.attrib["Id"]: node.attrib["Target"] for node in relationships}
        sheets = workbook.find(f"{{{NS_MAIN}}}sheets")
        if sheets is None:
            return result
        for sheet in sheets:
            relationship_id = sheet.attrib[f"{{{NS_REL}}}id"]
            target = targets[relationship_id].lstrip("/")
            xml_path = target if target.startswith("xl/") else posixpath.normpath(f"xl/{target}")
            rows: dict[int, list[str]] = {}
            for _event, element in ET.iterparse(io.BytesIO(archive.read(xml_path)), events=("end",)):
                if element.tag != f"{{{NS_MAIN}}}row":
                    continue
                row_values = []
                for cell in element.findall(f"{{{NS_MAIN}}}c"):
                    if column_number(cell.attrib.get("r", "")) > 64:
                        continue
                    cell_type = cell.attrib.get("t")
                    value_node = cell.find(f"{{{NS_MAIN}}}v")
                    if cell_type == "s" and value_node is not None and value_node.text is not None:
                        value = shared[int(value_node.text)]
                    elif cell_type == "inlineStr":
                        value = "".join(node.text or "" for node in cell.iter(f"{{{NS_MAIN}}}t"))
                    elif value_node is not None:
                        value = value_node.text or ""
                    else:
                        value = ""
                    if text(value):
                        row_values.append(text(value))
                if row_values:
                    rows[int(element.attrib["r"])] = row_values
                element.clear()
            result[sheet.attrib["name"]] = rows
    return result


def section_contexts(path: Path) -> dict[str, list[tuple[int, str]]]:
    if path.suffix.lower() != ".xlsx":
        return {}
    output: dict[str, list[tuple[int, str]]] = {}
    for sheet, rows in sparse_xlsx_rows(path).items():
        title_rows = {
            row_no: " | ".join(values)
            for row_no, values in rows.items()
            if len(values) == 1
            and len(values[0]) <= 200
            and "BOM" in values[0].upper()
            and not values[0].upper().startswith("BOM功能描述")
        }
        headers = sorted(
            row_no
            for row_no, values in rows.items()
            if len({field for value in values if (field := generic_header_field(value))}) >= 3
        )
        sections: list[tuple[int, str]] = []
        previous_header = 0
        for header in headers:
            nearby = [(row_no, values) for row_no, values in rows.items() if max(previous_header + 1, header - 8) <= row_no < header]
            nearby.sort()
            preferred_indexes = [index for index, (row_no, _values) in enumerate(nearby) if row_no in title_rows]
            title_indexes = [
                index
                for index, (_row, values) in enumerate(nearby)
                if "BOM" in " | ".join(values).upper() or "物料清单" in " | ".join(values)
            ]
            if preferred_indexes:
                selected = nearby[preferred_indexes[-1] :]
            elif title_indexes:
                selected = nearby[title_indexes[-1] :]
            elif nearby:
                selected = nearby[-1:]
            else:
                selected = []
            context_text = " / ".join(" | ".join(values) for _row, values in selected)[:1000]
            sections.append((header, context_text))
            previous_header = header
        for title_row, title_value in title_rows.items():
            sections.append((title_row, title_value[:1000]))
        sections.sort(key=lambda item: item[0])
        if not sections and rows:
            first_row = min(rows)
            sections.append((first_row, " | ".join(rows[first_row])[:1000]))
        output[sheet] = sections
    return output


def section_title(sections: dict[str, list[tuple[int, str]]], sheet: str, source_row: int) -> str:
    candidates = [title for header, title in sections.get(sheet, []) if header < source_row]
    return candidates[-1] if candidates else ""


def derive_context(filename: str, sheet: str, title: str) -> Context:
    project = project_for(filename)
    combined = f"{title} {sheet} {filename}"
    upper = unicodedata.normalize("NFKC", combined).upper()
    board_patterns = (
        ("PSSENSOR", "光感小板"),
        ("SUBLCM", "屏小板"),
        ("SIM-FPC", "SIM FPC"),
        ("TYPE-C", "TYPE-C小板"),
        ("TYPEC", "TYPE-C小板"),
        ("MOTOR", "马达小板"),
        ("WIFI", "WiFi小板"),
        ("DMR&RTK", "DMR/RTK小板"),
        ("DMR", "DMR/LTE小板"),
        ("N102", "N102小板"),
        ("座充", "座充指示小板"),
        ("_CH_", "座充指示小板"),
        ("ANT", "天线小板"),
        ("5G小板", "5G天线小板"),
        ("SUB", "SUB小板"),
    )
    board_type = next((label for marker, label in board_patterns if marker in upper), "")
    if not board_type:
        board_type = sheet if normalized_header(sheet) not in {"bom", "sheet1", "sheet2"} else f"{project}项目板"

    internal_model = ""
    version_match = re.search(r"PCB\s*版本\s*[:：]?\s*([A-Z0-9_.-]+)", upper)
    if version_match:
        internal_model = version_match.group(1).strip("-_.")
    model_matches = re.findall(r"(?<![A-Z0-9])([0-9](?:SD|SH|PF|P)[0-9]{4,6}[A-Z]?)(?![A-Z0-9])", upper)
    if model_matches:
        internal_model = model_matches[-1]
    if not internal_model and project == "J587":
        match = re.search(r"J587[_-]SUBA\d+[_-]V\d+", Path(filename).stem.upper())
        internal_model = match.group(0).replace("-", "_") if match else ""
    return Context(project=project, board_type=board_type[:80], internal_model=internal_model[:80], title=title)


def translated_reasons(codes: Iterable[str]) -> str:
    unique = []
    for code in codes:
        code = text(code)
        if code and code not in unique:
            unique.append(code)
    return "；".join(REASON_TEXT.get(code, code) for code in unique)


def review_action(codes: Iterable[str]) -> str:
    code_set = set(codes)
    if "STABLE_IDENTITY_CONFLICT" in code_set:
        return "核对完整 MPN、封装和关键规格；冲突消除前保持分离"
    if "STABLE_MATERIAL_IDENTITY_MISSING" in code_set:
        return "补齐完整 MPN 或该品类全部必需规格后再匹配"
    if "CATEGORY_OR_COUNT_UNIT_NOT_DETERMINISTIC" in code_set:
        return "确认物料类别、计量单位及该品类必需规格"
    if "BOM_QUANTITY_NOT_VALID" in code_set:
        return "确认单机用量和单位；不得把使用次数或范围位号当数量"
    if "TEMPLATE_STRICT_MATCH_AMBIGUOUS" in code_set:
        return "补齐品牌、MPN、封装和关键规格以消除多编码歧义"
    if "TEMPLATE_STRICT_MATCH_MISSING" in code_set:
        return "补齐稳定身份并与来源表逐行核对后再关联正式编码"
    if "SOURCE_CONTEXT_VERSION_CONFLICT" in code_set:
        return "由工程负责人确认文件名、表内标题和实际板卡版本"
    return "人工核对来源证据后处理"


def exact_template_matches(template_row: TemplateRow, source_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    key = strict_key(template_row.specification)
    if not key:
        return []
    return [row for row in source_rows if key in source_identity_keys(row)]


def load_guarded_evidence(
    profile_csv: Path,
    classification_csv: Path,
    payload_json: Path,
    manifest_json: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, Any], dict[str, Any]]:
    profile_rows = read_csv_rows(profile_csv)
    classification_rows = read_csv_rows(classification_csv)
    payload = json.loads(payload_json.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_json.read_text(encoding="utf-8"))
    profile_by_ref = {row["source_ref"]: row for row in profile_rows}
    classification_by_ref = {row["source_ref"]: row for row in classification_rows}
    if len(profile_by_ref) != len(profile_rows) or len(classification_by_ref) != len(classification_rows):
        raise ValueError("DUPLICATE_SOURCE_REF")
    if set(profile_by_ref) != set(classification_by_ref):
        raise ValueError("EVIDENCE_SOURCE_REF_MISMATCH")
    for source_ref, profile in profile_by_ref.items():
        classification = classification_by_ref[source_ref]
        for field in ("file_sha256", "sheet", "source_row"):
            if profile[field] != classification[field]:
                raise ValueError(f"EVIDENCE_ROW_MISMATCH:{source_ref}:{field}")
    if payload.get("manifest_sha256") != manifest.get("manifest_sha256"):
        raise ValueError("PAYLOAD_MANIFEST_MISMATCH")
    codes = [text(item.get("internal_code")) for item in payload.get("materials", [])]
    stable_keys = [text(item.get("stable_key")) for item in payload.get("materials", [])]
    if not codes or len(codes) != len(set(codes)) or len(stable_keys) != len(set(stable_keys)):
        raise ValueError("PAYLOAD_MATERIAL_IDENTITY_INVALID")
    return profile_rows, classification_rows, payload, manifest


def build_workbook_data(
    source_dir: Path,
    template_path: Path,
    profile_rows: list[dict[str, str]],
    classification_rows: list[dict[str, str]],
    payload: dict[str, Any],
    template_rows: list[TemplateRow],
    source_manifest: dict[str, Any],
    template_sha256: str,
) -> dict[str, Any]:
    profile_by_ref = {row["source_ref"]: row for row in profile_rows}
    classification_by_ref = {row["source_ref"]: row for row in classification_rows}
    nonarchive_refs = {
        row["source_ref"] for row in classification_rows if row["material_classification"] != "ARCHIVE_ONLY"
    }
    review_refs = {row["source_ref"] for row in classification_rows if row["classification"] == "NEEDS_REVIEW"}

    contexts_by_file: dict[str, dict[str, list[tuple[int, str]]]] = {}
    for item in source_manifest["files"]:
        filename = text(item["filename"])
        path = source_dir / filename
        contexts_by_file[filename] = section_contexts(path)

    context_by_ref: dict[str, Context] = {}
    for row in profile_rows:
        source_row = int(row["source_row"])
        title = section_title(contexts_by_file.get(row["filename"], {}), row["sheet"], source_row)
        context_by_ref[row["source_ref"]] = derive_context(row["filename"], row["sheet"], title)

    a200_sources = [row for row in profile_rows if row["source_ref"] in nonarchive_refs and project_for(row["filename"]) == "A200"]
    template_matches: dict[int, list[dict[str, str]]] = {
        row.source_row: exact_template_matches(row, a200_sources) for row in template_rows
    }
    template_rows_by_key: dict[str, list[TemplateRow]] = defaultdict(list)
    for row in template_rows:
        template_rows_by_key[strict_key(row.specification)].append(row)

    source_to_template: dict[str, TemplateRow] = {}
    for source in a200_sources:
        matched_rows = {
            template.source_row: template
            for key in source_identity_keys(source)
            for template in template_rows_by_key.get(key, [])
        }
        if len(matched_rows) == 1:
            source_to_template[source["source_ref"]] = next(iter(matched_rows.values()))

    code_to_payload = {text(item["internal_code"]): item for item in payload["materials"]}
    stable_to_payload = {text(item["stable_key"]): item for item in payload["materials"]}
    template_resolution: dict[int, dict[str, str]] = {}
    formal_template_rows: dict[str, list[TemplateRow]] = defaultdict(list)
    template_candidates: list[dict[str, Any]] = []
    for template in template_rows:
        matched = template_matches[template.source_row]
        codes = {
            classification_by_ref[row["source_ref"]]["internal_code"]
            for row in matched
            if classification_by_ref[row["source_ref"]]["material_classification"] == "ELIGIBLE"
            and classification_by_ref[row["source_ref"]]["internal_code"]
        }
        if len(codes) == 1:
            code = next(iter(codes))
            payload_item = code_to_payload.get(code)
            if payload_item is None:
                raise ValueError(f"TEMPLATE_CODE_NOT_IN_PAYLOAD:{code}")
            template_resolution[template.source_row] = {
                "library_id": text(payload_item["stable_key"]),
                "internal_code": code,
                "status": "严格对应既有正式物料",
                "basis": "模板规格与来源规格/型号/完整 MPN 严格相等",
            }
            formal_template_rows[code].append(template)
        else:
            reason = "TEMPLATE_STRICT_MATCH_AMBIGUOUS" if len(codes) > 1 else "TEMPLATE_STRICT_MATCH_MISSING"
            candidate_id = stable_candidate_id("CAND-TPL", f"{template_sha256}:Sheet1:{template.source_row}")
            template_resolution[template.source_row] = {
                "library_id": candidate_id,
                "internal_code": "",
                "status": "模板物料待确认",
                "basis": REASON_TEXT[reason],
                "reason": reason,
            }
            template_candidates.append({"row": template, "candidate_id": candidate_id, "reason": reason})

    library_rows: list[list[Any]] = []
    for material in sorted(payload["materials"], key=lambda item: text(item["internal_code"])):
        source_records = material.get("source_rows", [])
        source_refs = [text(item.get("source_ref")) for item in source_records]
        projects = {context_by_ref[source_ref].project for source_ref in source_refs if source_ref in context_by_ref}
        filenames = {profile_by_ref[source_ref]["filename"] for source_ref in source_refs if source_ref in profile_by_ref}
        code = text(material["internal_code"])
        if formal_template_rows.get(code):
            projects.update(row.project for row in formal_template_rows[code])
            filenames.add(template_path.name)
        mpn_model = text(material.get("manufacturer_part_number")) or text(material.get("model"))
        specification = text(material.get("specification")) or text(material.get("model")) or text(material.get("manufacturer_part_number"))
        library_rows.append(
            [
                text(material["stable_key"]),
                code,
                "既有正式编码（离线证据）",
                text(material.get("category_name")) or text(material.get("category_code")),
                text(material.get("standard_name")),
                specification,
                text(material.get("manufacturer")),
                mpn_model,
                text(material.get("package")),
                "PCS",
                "、".join(sorted(projects)),
                len(filenames),
                len(source_refs) + len(formal_template_rows.get(code, [])),
                IDENTITY_TEXT.get(text(material.get("identity_method")), text(material.get("identity_method"))),
                "",
            ]
        )

    raw_candidate_by_ref: dict[str, str] = {}
    for classification in sorted(classification_rows, key=lambda row: (row.get("file_sha256", ""), row.get("sheet", ""), int(row.get("source_row", "0")))):
        if classification["material_classification"] != "NEEDS_REVIEW":
            continue
        source_ref = classification["source_ref"]
        source = profile_by_ref[source_ref]
        candidate_id = stable_candidate_id("CAND-SRC", source_ref)
        raw_candidate_by_ref[source_ref] = candidate_id
        codes = [code for code in classification["reason_codes"].split("|") if code]
        library_rows.append(
            [
                candidate_id,
                "",
                "待人工确认（未分配正式编码）",
                text(classification.get("category_code")),
                text(source.get("material_name")),
                source_description(source),
                text(source.get("manufacturer")),
                text(source.get("manufacturer_part_no")) or text(source.get("model")),
                text(source.get("package")),
                "",
                context_by_ref[source_ref].project,
                1,
                1,
                "来源行隔离；尚无安全归并依据",
                translated_reasons(codes),
            ]
        )

    for candidate in template_candidates:
        template = candidate["row"]
        library_rows.append(
            [
                candidate["candidate_id"],
                "",
                "模板物料待人工确认（未分配正式编码）",
                "",
                template.specification,
                template.specification,
                template.brand,
                "",
                "",
                "",
                template.project,
                1,
                1,
                "模板来源行隔离；尚无安全归并依据",
                REASON_TEXT[candidate["reason"]],
            ]
        )

    standard_entries: list[dict[str, Any]] = []
    for template in template_rows:
        resolution = template_resolution[template.source_row]
        linked_review = any(
            classification_by_ref[source["source_ref"]]["classification"] == "NEEDS_REVIEW"
            for source in template_matches[template.source_row]
        )
        template_status = resolution["status"] + ("；关联来源另有待确认" if linked_review else "")
        standard_entries.append(
            {
                "entry_id": f"template:{template.source_row}",
                "sort": (0, template.source_row),
                "project": template.project,
                "board_type": template.board_type,
                "internal_model": template.internal_model,
                "specification": template.specification,
                "brand": template.brand,
                "quantity": template.quantity,
                "substitute": template.substitute,
                "supplier": template.supplier,
                "order_quantity": template.order_quantity,
                "inventory": template.inventory,
                "status": template_status,
                "record_id": resolution["library_id"],
                "internal_code": resolution["internal_code"],
                "source_label": f"{template_path.name}/Sheet1/{template.source_row}",
            }
        )

    project_order = {name: index for index, name in enumerate(("A200", "1928C", "A118", "G20-G15G", "J587", "V700"))}
    source_to_entry: dict[str, str] = {}
    for source in profile_rows:
        source_ref = source["source_ref"]
        if source_ref not in nonarchive_refs:
            continue
        if source_ref in source_to_template:
            source_to_entry[source_ref] = f"template:{source_to_template[source_ref].source_row}"
            continue
        context = context_by_ref[source_ref]
        classification = classification_by_ref[source_ref]
        entry_id = f"source:{source_ref}"
        source_to_entry[source_ref] = entry_id
        quantity = excel_number(source.get("quantity"), positive=True)
        if quantity is None:
            quantity = explicit_reference_quantity(source.get("reference"))
        record_id = classification.get("internal_stable_id") or raw_candidate_by_ref.get(source_ref, "")
        if classification["material_classification"] != "ELIGIBLE":
            status = "原始来源待确认"
        elif classification["classification"] == "NEEDS_REVIEW":
            status = "物料已对应既有正式编码；BOM数量/结构待确认"
        else:
            status = "原始来源已对应既有正式物料"
        standard_entries.append(
            {
                "entry_id": entry_id,
                "sort": (1, source["filename"], source["sheet"], int(source["source_row"])),
                "project": context.project,
                "board_type": context.board_type,
                "internal_model": context.internal_model or "待确认",
                "specification": source_description(source),
                "brand": source.get("manufacturer", ""),
                "quantity": quantity,
                "substitute": source.get("substitute", ""),
                "supplier": "",
                "order_quantity": None,
                "inventory": None,
                "status": status,
                "record_id": record_id,
                "internal_code": classification.get("internal_code", ""),
                "source_label": f"{source['filename']}/{source['sheet']}/{source['source_row']}",
            }
        )

    standard_entries.sort(
        key=lambda item: (
            project_order.get(item["project"], 99),
            item["project"],
            item["sort"],
        )
    )
    entry_sequence: dict[str, int] = {}
    standard_rows: list[list[Any]] = []
    standard_comments: dict[int, str] = {}
    for sequence, entry in enumerate(standard_entries, 1):
        entry_sequence[entry["entry_id"]] = sequence
        standard_rows.append(
            [
                sequence,
                entry["project"],
                entry["board_type"],
                entry["internal_model"],
                entry["specification"],
                entry["brand"],
                entry["quantity"],
                entry["substitute"],
                entry["supplier"],
                entry["order_quantity"],
                None,
                None,
                entry["inventory"],
            ]
        )
        code_or_id = entry["internal_code"] or entry["record_id"]
        standard_comments[sequence + 1] = f"状态：{entry['status']}\n记录：{code_or_id}\n来源：{entry['source_label']}"

    source_mapping_rows: list[list[Any]] = []
    for source in profile_rows:
        source_ref = source["source_ref"]
        if source_ref not in nonarchive_refs:
            continue
        classification = classification_by_ref[source_ref]
        formal = classification["material_classification"] == "ELIGIBLE"
        library_id = classification.get("internal_stable_id") if formal else raw_candidate_by_ref[source_ref]
        source_mapping_rows.append(
            [
                library_id,
                classification.get("internal_code", "") if formal else "",
                source_ref,
                "原始来源",
                context_by_ref[source_ref].project,
                source["filename"],
                source["sheet"],
                int(source["source_row"]),
                source["file_sha256"][:12],
                "已对应既有正式物料" if formal else "待确认隔离",
                IDENTITY_TEXT.get(classification.get("mapping_method", ""), classification.get("mapping_method", "")),
                entry_sequence[source_to_entry[source_ref]],
            ]
        )
    for template in template_rows:
        resolution = template_resolution[template.source_row]
        source_mapping_rows.append(
            [
                resolution["library_id"],
                resolution["internal_code"],
                f"template:{template_sha256[:16]}:Sheet1:{template.source_row}",
                "人工整理模板",
                template.project,
                template_path.name,
                "Sheet1",
                template.source_row,
                template_sha256[:12],
                resolution["status"],
                resolution["basis"],
                entry_sequence[f"template:{template.source_row}"],
            ]
        )

    review_rows: list[list[Any]] = []
    for classification in classification_rows:
        if classification["classification"] != "NEEDS_REVIEW":
            continue
        source_ref = classification["source_ref"]
        source = profile_by_ref[source_ref]
        codes = [code for code in classification["reason_codes"].split("|") if code]
        level = "物料身份" if classification["material_classification"] == "NEEDS_REVIEW" else "BOM数量/结构"
        review_rows.append(
            [
                stable_candidate_id("REV-SRC", source_ref),
                source_ref,
                context_by_ref[source_ref].project,
                level,
                translated_reasons(codes),
                review_action(codes),
                classification.get("internal_code", ""),
                source_description(source),
                source.get("manufacturer", ""),
                source["filename"],
                source["sheet"],
                int(source["source_row"]),
            ]
        )
    for candidate in template_candidates:
        template = candidate["row"]
        reason = candidate["reason"]
        review_rows.append(
            [
                stable_candidate_id("REV-TPL", f"{template_sha256}:Sheet1:{template.source_row}"),
                f"template:{template_sha256[:16]}:Sheet1:{template.source_row}",
                template.project,
                "模板与主数据关联",
                REASON_TEXT[reason],
                review_action([reason]),
                "",
                template.specification,
                template.brand,
                template_path.name,
                "Sheet1",
                template.source_row,
            ]
        )

    version_conflicts = []
    for filename, sheet_contexts in contexts_by_file.items():
        file_marker = re.search(r"([A-Z0-9]+[_-]SUBA\d+[_-]V\d+)", Path(filename).stem.upper())
        if not file_marker:
            continue
        title_markers = {
            match.group(1).replace("-", "_")
            for sections in sheet_contexts.values()
            for _header, title in sections
            for match in re.finditer(r"([A-Z0-9]+[_-]SUBA\d+[_-]V\d+)", title.upper())
        }
        expected_marker = file_marker.group(1).replace("-", "_")
        if title_markers and title_markers != {expected_marker}:
            version_conflicts.append(filename)
            review_rows.append(
                [
                    stable_candidate_id("REV-CTX", filename),
                    f"context:{sha256_path(source_dir / filename)[:16]}",
                    project_for(filename),
                    "来源上下文",
                    REASON_TEXT["SOURCE_CONTEXT_VERSION_CONFLICT"],
                    review_action(["SOURCE_CONTEXT_VERSION_CONFLICT"]),
                    "",
                    "",
                    "",
                    filename,
                    "",
                    "",
                ]
            )

    counts = {
        "source_rows": len(profile_rows),
        "archive_source_rows": sum(1 for row in classification_rows if row["material_classification"] == "ARCHIVE_ONLY"),
        "nonarchive_material_source_rows": len(nonarchive_refs),
        "classification_review_rows": len(review_refs),
        "material_review_source_rows": sum(1 for row in classification_rows if row["material_classification"] == "NEEDS_REVIEW"),
        "bom_review_source_rows": sum(1 for row in classification_rows if row["bom_classification"] == "NEEDS_REVIEW"),
        "formal_materials": len(payload["materials"]),
        "template_material_rows": len(template_rows),
        "template_rows_strictly_mapped": len(template_rows) - len(template_candidates),
        "template_rows_needing_review": len(template_candidates),
        "library_rows": len(library_rows),
        "standard_detail_rows": len(standard_rows),
        "review_rows": len(review_rows),
        "source_mapping_rows": len(source_mapping_rows),
        "source_context_version_conflicts": len(version_conflicts),
    }
    expected_source_mapping_count = len(nonarchive_refs) + len(template_rows)
    if len(source_mapping_rows) != expected_source_mapping_count:
        raise ValueError("SOURCE_MAPPING_COVERAGE_MISMATCH")
    if set(source_to_entry) != nonarchive_refs:
        raise ValueError("STANDARD_DETAIL_SOURCE_COVERAGE_MISMATCH")
    if review_refs - {row[1] for row in review_rows}:
        raise ValueError("REVIEW_SOURCE_COVERAGE_MISMATCH")

    return {
        "library_rows": library_rows,
        "standard_rows": standard_rows,
        "standard_comments": standard_comments,
        "review_rows": review_rows,
        "source_mapping_rows": source_mapping_rows,
        "counts": counts,
        "nonarchive_refs": sorted(nonarchive_refs),
        "review_refs": sorted(review_refs),
    }


HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
FORMAL_FILL = PatternFill("solid", fgColor="E2F0D9")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_GREY = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def style_tabular_sheet(sheet: Any, widths: list[int]) -> None:
    for cell in sheet[1]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(CELL_BORDER)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = copy(CELL_BORDER)


def write_rows(sheet: Any, headers: tuple[str, ...], rows: list[list[Any]]) -> None:
    sheet.append(list(headers))
    for row in rows:
        identifier_headers = {"内部记录ID", "正式ERP编码", "待确认ID", "来源引用", "来源SHA前12位"}
        cleaned = [
            value
            if isinstance(value, (int, float)) or value is None
            else safe_excel_text(value, redact_phone=headers[index] not in identifier_headers)
            for index, value in enumerate(row)
        ]
        sheet.append(cleaned)


def write_workbook(
    output: Path,
    workbook_data: dict[str, Any],
    source_manifest: dict[str, Any],
    template_path: Path,
    template_sha256: str,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "晨亿达 ERP"
    workbook.properties.title = "内部物料库"
    workbook.properties.subject = TASK
    workbook.properties.description = "按既有模板离线整理；正式编码只沿用既有导入证据。"
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass

    library = workbook.create_sheet("内部物料库")
    write_rows(library, LIBRARY_HEADERS, workbook_data["library_rows"])
    style_tabular_sheet(library, [28, 23, 27, 16, 28, 48, 22, 30, 14, 10, 22, 12, 12, 26, 48])
    for row_no in range(2, library.max_row + 1):
        status = text(library.cell(row_no, 3).value)
        fill = REVIEW_FILL if "待" in status else FORMAL_FILL
        library.cell(row_no, 3).fill = copy(fill)

    details = workbook.create_sheet("标准BOM明细")
    write_rows(details, TEMPLATE_HEADERS, workbook_data["standard_rows"])
    for row_no in range(2, details.max_row + 1):
        details.cell(row_no, 11).value = f'=IF(OR(G{row_no}="",J{row_no}=""),"",G{row_no}*J{row_no})'
        details.cell(row_no, 12).value = f'=IF(K{row_no}="","",MAX(K{row_no}-M{row_no},0))'
        for column in (7, 10, 11, 12, 13):
            details.cell(row_no, column).number_format = "0.######"
        comment = workbook_data["standard_comments"].get(row_no)
        if comment:
            details.cell(row_no, 1).comment = Comment(
                safe_excel_text(comment, limit=2000, redact_phone=False), "晨亿达 ERP"
            )
            if "待确认" in comment:
                details.cell(row_no, 1).fill = copy(REVIEW_FILL)
    style_tabular_sheet(details, [9, 14, 18, 20, 55, 20, 11, 24, 24, 12, 13, 13, 11])
    validation = DataValidation(type="decimal", operator="between", formula1="0", formula2="999999999", allow_blank=True)
    validation.error = "请输入非负数量"
    validation.errorTitle = "数量无效"
    validation.prompt = "未知数量请留空，不要填文字或猜测值。"
    validation.promptTitle = "数量填写规则"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    details.add_data_validation(validation)
    for column in ("G", "J", "M"):
        validation.add(f"{column}2:{column}{max(2, details.max_row)}")

    review = workbook.create_sheet("待确认")
    write_rows(review, REVIEW_HEADERS, workbook_data["review_rows"])
    style_tabular_sheet(review, [22, 28, 15, 18, 45, 46, 22, 52, 20, 32, 24, 12])
    for row_no in range(2, review.max_row + 1):
        review.cell(row_no, 4).fill = copy(ERROR_FILL)

    source = workbook.create_sheet("来源映射")
    write_rows(source, SOURCE_HEADERS, workbook_data["source_mapping_rows"])
    style_tabular_sheet(source, [28, 23, 28, 16, 15, 32, 24, 12, 16, 26, 34, 15])

    notes = workbook.create_sheet("来源与说明")
    notes.sheet_view.showGridLines = False
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 72
    notes.column_dimensions["C"].width = 72
    notes.column_dimensions["D"].width = 18
    notes["A1"] = "内部物料库整理说明"
    notes["A1"].font = Font(size=16, bold=True, color="17365D")
    notes.merge_cells("A1:D1")
    summary = [
        ("任务编号", TASK),
        ("生成时间", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("模板文件", template_path.name),
        ("模板工作表", "Sheet1（第二张，整理后标准）；原BOM（第一张）仅作对照"),
        ("模板 SHA-256", template_sha256),
        ("模板标准列", "、".join(TEMPLATE_HEADERS)),
        ("既有正式物料", workbook_data["counts"]["formal_materials"]),
        ("内部物料库总行", workbook_data["counts"]["library_rows"]),
        ("标准BOM明细", workbook_data["counts"]["standard_detail_rows"]),
        ("来源分类待确认", workbook_data["counts"]["classification_review_rows"]),
        ("模板关联待确认", workbook_data["counts"]["template_rows_needing_review"]),
        ("数据库影响", "无；本工作簿未导入 PostgreSQL、SQLite 或 D1"),
    ]
    row_no = 3
    for label, value in summary:
        notes.cell(row_no, 1).value = label
        notes.cell(row_no, 2).value = safe_excel_text(value) if not isinstance(value, (int, float)) else value
        notes.cell(row_no, 1).fill = copy(SUBHEADER_FILL)
        notes.cell(row_no, 1).font = Font(bold=True)
        notes.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=4)
        row_no += 1
    row_no += 1
    notes.cell(row_no, 1).value = "来源文件"
    notes.cell(row_no, 1).font = Font(size=12, bold=True, color="17365D")
    row_no += 1
    source_header_row = row_no
    for column, value in enumerate(("文件名", "用途", "SHA-256", "字节数"), 1):
        cell = notes.cell(row_no, column)
        cell.value = value
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
    row_no += 1
    for item in source_manifest["files"]:
        notes.cell(row_no, 1).value = safe_excel_text(item["filename"])
        notes.cell(row_no, 2).value = "来源表（只读）"
        notes.cell(row_no, 3).value = item["sha256"]
        notes.cell(row_no, 4).value = int(item["size_bytes"])
        row_no += 1
    notes.cell(row_no, 1).value = template_path.name
    notes.cell(row_no, 2).value = "人工整理模板（只读）"
    notes.cell(row_no, 3).value = template_sha256
    notes.cell(row_no, 4).value = template_path.stat().st_size
    row_no += 2
    notes.cell(row_no, 1).value = "处理边界"
    notes.cell(row_no, 1).font = Font(size=12, bold=True, color="17365D")
    row_no += 1
    boundaries = (
        "正式 ERP 编码只沿用 SELFHOST-LANDING-TASK02 已存在的映射证据，本任务不生成新正式编码。",
        "缺规格、身份冲突、类别/单位不确定或数量不可靠的来源均进入“待确认”，不模糊合并。",
        "供应商字段只使用人工模板中的明确值；生产厂商/品牌不冒充供应商。",
        "替代料只保留为来源证据，不自动形成正式替代关系。",
        "自由备注和联系电话不进入结果；所有外来文本按文本写入，避免公式注入。",
        "来源无法可靠提取内部型号时明确显示“待确认”，不会猜测板卡编码。",
        "本工作簿是离线整理结果，不代表已完成审核、数据库入库或生产生效。",
    )
    for index, boundary in enumerate(boundaries, 1):
        notes.cell(row_no, 1).value = index
        notes.cell(row_no, 2).value = boundary
        notes.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=4)
        row_no += 1
    for row in notes.iter_rows(min_row=3, max_row=row_no - 1, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = copy(CELL_BORDER)
    notes.freeze_panes = f"A{source_header_row + 1}"

    output.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{output.stem}-", suffix=".xlsx", dir=output.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        workbook.save(temporary)
        temporary.chmod(0o600)
        os.replace(temporary, output)
        output.chmod(0o600)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def validate_workbook(output: Path, data: dict[str, Any]) -> dict[str, Any]:
    expected_sheets = ["内部物料库", "标准BOM明细", "待确认", "来源映射", "来源与说明"]
    with ZipFile(output) as archive:
        names = archive.namelist()
        macro_entries = [name for name in names if "vbaproject" in name.lower()]
        external_entries = [name for name in names if name.startswith("xl/externalLinks/")]
    if macro_entries or external_entries:
        raise ValueError("OUTPUT_ACTIVE_CONTENT_FORBIDDEN")

    workbook = openpyxl.load_workbook(output, read_only=False, data_only=False, keep_links=False)
    try:
        if workbook.sheetnames != expected_sheets:
            raise ValueError("OUTPUT_SHEET_SET_MISMATCH")
        expected_headers = {
            "内部物料库": LIBRARY_HEADERS,
            "标准BOM明细": TEMPLATE_HEADERS,
            "待确认": REVIEW_HEADERS,
            "来源映射": SOURCE_HEADERS,
        }
        expected_rows = {
            "内部物料库": data["counts"]["library_rows"],
            "标准BOM明细": data["counts"]["standard_detail_rows"],
            "待确认": data["counts"]["review_rows"],
            "来源映射": data["counts"]["source_mapping_rows"],
        }
        for name, headers in expected_headers.items():
            sheet = workbook[name]
            actual_headers = tuple(text(sheet.cell(1, column).value) for column in range(1, len(headers) + 1))
            if actual_headers != headers or sheet.max_row - 1 != expected_rows[name]:
                raise ValueError(f"OUTPUT_TABULAR_CONTRACT_MISMATCH:{name}")

        formulas = []
        phone_like_cells = []
        sensitive_cells = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type == "f":
                        formulas.append((sheet.title, cell.coordinate, value))
                    if isinstance(value, str):
                        identifier_like = bool(
                            re.fullmatch(r"[0-9a-f]{12,64}", value, re.I)
                            or re.fullmatch(r"material:[0-9a-f]+", value, re.I)
                            or re.fullmatch(r"(?:CAND|REV)-(?:SRC|TPL|CTX)-[A-F0-9]+", value)
                            or value.startswith(("template:", "context:"))
                        )
                        if PHONE_LIKE_RE.search(value) and not identifier_like:
                            phone_like_cells.append((sheet.title, cell.coordinate))
                        if SENSITIVE_RE.search(value):
                            sensitive_cells.append((sheet.title, cell.coordinate))
                    if cell.comment and (PHONE_LIKE_RE.search(cell.comment.text) or SENSITIVE_RE.search(cell.comment.text)):
                        sensitive_cells.append((sheet.title, f"comment:{cell.coordinate}"))
        expected_formula_count = data["counts"]["standard_detail_rows"] * 2
        if len(formulas) != expected_formula_count:
            raise ValueError("OUTPUT_FORMULA_COUNT_MISMATCH")
        for sheet_name, coordinate, formula in formulas:
            row_no = int(re.search(r"\d+", coordinate).group())
            expected = {
                f"K{row_no}": f'=IF(OR(G{row_no}="",J{row_no}=""),"",G{row_no}*J{row_no})',
                f"L{row_no}": f'=IF(K{row_no}="","",MAX(K{row_no}-M{row_no},0))',
            }
            if sheet_name != "标准BOM明细" or expected.get(coordinate) != formula:
                raise ValueError(f"OUTPUT_UNEXPECTED_FORMULA:{sheet_name}:{coordinate}")
        if phone_like_cells or sensitive_cells:
            raise ValueError("OUTPUT_SENSITIVE_CONTENT_DETECTED")

        library = workbook["内部物料库"]
        formal_codes = []
        for row_no in range(2, library.max_row + 1):
            code = text(library.cell(row_no, 2).value)
            status = text(library.cell(row_no, 3).value)
            if code:
                formal_codes.append(code)
            if "待" in status and code:
                raise ValueError("CANDIDATE_RECEIVED_FORMAL_CODE")
        if len(formal_codes) != data["counts"]["formal_materials"] or len(formal_codes) != len(set(formal_codes)):
            raise ValueError("FORMAL_CODE_UNIQUENESS_MISMATCH")

        mappings = workbook["来源映射"]
        mapped_refs = {text(mappings.cell(row_no, 3).value) for row_no in range(2, mappings.max_row + 1)}
        if set(data["nonarchive_refs"]) - mapped_refs:
            raise ValueError("OUTPUT_SOURCE_MAPPING_COVERAGE_MISMATCH")
        reviews = workbook["待确认"]
        review_refs = {text(reviews.cell(row_no, 2).value) for row_no in range(2, reviews.max_row + 1)}
        if set(data["review_refs"]) - review_refs:
            raise ValueError("OUTPUT_REVIEW_COVERAGE_MISMATCH")
    finally:
        workbook.close()
    return {
        "sheet_contract": "PASS",
        "source_mapping_coverage": "PASS",
        "review_coverage": "PASS",
        "formal_code_uniqueness": "PASS",
        "candidate_formal_code_guard": "PASS",
        "formula_contract": "PASS",
        "formula_cells": expected_formula_count,
        "macro_entries": 0,
        "external_link_entries": 0,
        "phone_like_cells": 0,
        "sensitive_cells": 0,
    }


def atomic_write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".json", dir=path.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        temporary.chmod(0o600)
        os.replace(temporary, path)
        path.chmod(0o600)
    finally:
        if temporary.exists():
            temporary.unlink()


def export_library(
    *,
    source_dir: Path,
    template_path: Path,
    profile_csv: Path,
    classification_csv: Path,
    payload_json: Path,
    manifest_json: Path,
    output: Path,
    report_path: Path,
    confirmation: str,
) -> dict[str, Any]:
    if confirmation != CONFIRMATION:
        raise ValueError("CONFIRMATION_REQUIRED")
    if output.suffix.lower() != ".xlsx" or output.resolve() == template_path.resolve():
        raise ValueError("OUTPUT_PATH_INVALID")
    os.umask(0o077)
    profile_rows, classification_rows, payload, source_manifest = load_guarded_evidence(
        profile_csv, classification_csv, payload_json, manifest_json
    )
    source_before = verify_manifest(source_dir, source_manifest)
    template_before = file_snapshot(template_path)
    template_rows = parse_template(template_path)
    workbook_data = build_workbook_data(
        source_dir,
        template_path,
        profile_rows,
        classification_rows,
        payload,
        template_rows,
        source_manifest,
        template_before["sha256"],
    )
    write_workbook(output, workbook_data, source_manifest, template_path, template_before["sha256"])
    validation = validate_workbook(output, workbook_data)
    source_after = verify_manifest(source_dir, source_manifest)
    template_after = file_snapshot(template_path)
    if source_before != source_after or template_before != template_after:
        raise ValueError("INPUT_CHANGED_DURING_EXPORT")
    mode = stat.S_IMODE(output.stat().st_mode)
    if mode != 0o600:
        raise ValueError("OUTPUT_MODE_INVALID")
    report = {
        "task": TASK,
        "result": "OFFLINE_INTERNAL_MATERIAL_LIBRARY_CREATED_REVIEW_REQUIRED",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_manifest_sha256": source_manifest["manifest_sha256"],
        "template": {
            "filename": template_path.name,
            "sha256": template_before["sha256"],
            "reference_sheet": "原BOM",
            "standard_sheet": "Sheet1",
            "headers": list(TEMPLATE_HEADERS),
            "material_rows": len(template_rows),
        },
        "output": {
            "path": str(output),
            "sha256": sha256_path(output),
            "size_bytes": output.stat().st_size,
            "mode": oct(mode),
        },
        "counts": workbook_data["counts"],
        "validation": validation | {"source_files_unchanged": "PASS", "template_unchanged": "PASS"},
        "database_writes": 0,
        "deployment_changes": 0,
    }
    atomic_write_json(report_path, report)
    return report


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--profile-csv", type=Path, required=True)
    parser.add_argument("--classification-csv", type=Path, required=True)
    parser.add_argument("--payload-json", type=Path, required=True)
    parser.add_argument("--manifest-json", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--confirm", required=True)
    args = parser.parse_args()
    report = export_library(
        source_dir=args.source_dir,
        template_path=args.template,
        profile_csv=args.profile_csv,
        classification_csv=args.classification_csv,
        payload_json=args.payload_json,
        manifest_json=args.manifest_json,
        output=args.output,
        report_path=args.report,
        confirmation=args.confirm,
    )
    print(
        json.dumps(
            {
                "result": report["result"],
                "output": report["output"],
                "counts": report["counts"],
                "validation": report["validation"],
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
